#!/usr/bin/env python3
#
# ex:ts=4:sw=4:sts=4:et
# -*- tab-width: 4; c-basic-offset: 4; indent-tabs-mode: nil -*-
#
# Security Response Tool Commandline Tool
#
# Copyright (C) 2021-2022  Wind River Systems
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License version 2 as
# published by the Free Software Foundation.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License along
# with this program; if not, write to the Free Software Foundation, Inc.,
# 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.

### Usage Examples (run from top level directory)
# Init WR products:   ./bin/wrsrtool_wr.py --init-products
import shlex
import os
import sys
import argparse
import json
import subprocess
import time
import pytz
import csv
import os
import re
import shlex
import logging
import json
from collections import Counter
from datetime import datetime, date, timedelta
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import requests
import logging
from django.db.models import Q, F
from django.db import Error
from django.db.models import Subquery

# load the srt.sqlite schema indexes
dir_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.insert(0, dir_path)
from common.srt_schema import ORM
from common.srtool_sql import *
from common.srtool_progress import *
from common.srtool_common import log_error

# Setup:
logger = logging.getLogger("srt")

SRT_BASE_DIR = os.environ.get('SRT_BASE_DIR','.')
SRT_REPORT_DIR = f"{SRT_BASE_DIR}/reports"
SRTOOL_DEST_DB_PATH = f"{SRT_BASE_DIR}/srt_wrlinux.sqlite"

verbose = False
test = False
cmd_count = 0
cmd_skip = 0
force_update = False
srt_lx_remote = ''
srt_lx_path = ''

User_SRTool_ID = 2
User_All_ID = 3
Group_Reader_ID = 1

#################################
# Helper methods
#

def debugMsg(msg):
    if verbose:
        print(msg)

overrides = {}

def set_override(key,value=None):
    if not value is None:
        overrides[key] = value
    elif key in os.environ.keys():
        overrides[key] = 'yes' if os.environ[key].startswith('1') else 'no'
    else:
        overrides[key] = 'no'
    if 'yes' == overrides[key]:
        print("OVERRIDE: %s = %s" % (key,overrides[key]))

def get_override(key):
    if key in overrides.keys():
        return 'yes' == overrides[key]
    return False

# quick development/debugging support
def _log(msg):
    DBG_LVL =  os.environ['SRTDBG_LVL'] if ('SRTDBG_LVL' in os.environ) else 2
    DBG_LOG =  os.environ['SRTDBG_LOG'] if ('SRTDBG_LOG' in os.environ) else '/tmp/srt_dbg.log'
    if 1 == DBG_LVL:
        print(msg)
    elif 2 == DBG_LVL:
        f1=open(DBG_LOG, 'a')
        f1.write("|" + msg + "|\n" )
        f1.close()

def get_tag_key(tag,key,default=None):
    d = json.loads(tag)
    if key in d:
        return d[key]
    return default

# Sub Process calls
def execute_process(*args):
    cmd_list = []
    for arg in args:
        if not arg: continue
        if isinstance(arg, (list, tuple)):
            # Flatten all the way down
            for a in arg:
                if not a: continue
                cmd_list.append(a)
        else:
            cmd_list.append(arg)

    if verbose: print(f"EXECUTE_PROCESS:{cmd_list}:PWD={os.getcwd()}")
    result = subprocess.run(cmd_list, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return result.returncode,result.stdout.decode('utf-8'),result.stderr.decode('utf-8')


def srtsetting_get(conn,key,default_value,is_dict=True):
    cur = SQL_CURSOR(conn)
    # Fetch the key for SrtSetting
    sql = f"""SELECT * FROM orm_srtsetting WHERE `name` = ?"""
    try:
        srtsetting = SQL_EXECUTE(cur, sql,(key,)).fetchone()
        if srtsetting:
            if is_dict:
                return(srtsetting['value'])
            else:
                return(srtsetting[ORM.SRTSETTING_VALUE])
    except Exception as e:
        print("ERROR:%s" % (e))
    return(default_value)

def srtsetting_set(conn,key,value,is_dict=True):
    cur = SQL_CURSOR(conn)
    # Set the key value for SrtSetting
    sql = f"""SELECT * FROM orm_srtsetting WHERE `name` = ?"""
    srtsetting = SQL_EXECUTE(cur, sql,(key,)).fetchone()
    if not srtsetting:
        sql = ''' INSERT INTO orm_srtsetting (name, helptext, value) VALUES (?,?,?)'''
        SQL_EXECUTE(cur, sql, (key,'',value))
        print("INSERT:%s:%s:" % (key,value))
    else:
        print("UPDATE[%d]:%s:%s:" % (srtsetting[ORM.SRTSETTING_ID],key,value))
        sql = ''' UPDATE orm_srtsetting
                  SET value=?
                  WHERE id=?'''
        if is_dict:
            SQL_EXECUTE(cur, sql, (value,srtsetting['id']))
        else:
            SQL_EXECUTE(cur, sql, (value,srtsetting[ORM.SRTSETTING_ID]))
    SQL_COMMIT(conn)

###############################################################################
# Excel/openpyxl common look and feel formatting objects

pyxl_thin = Side(border_style="thin")
pyxl_double = Side(border_style="double")
pyxl_border_left = Border(left=pyxl_thin)
pyxl_border_bottom = Border(bottom=pyxl_thin)
pyxl_border_bottom_left = Border(bottom=pyxl_thin, left=pyxl_thin)
pyxl_alignment_left = Alignment(horizontal='left')
pyxl_alignment_right = Alignment(horizontal='right')
pyxl_alignment_top = Alignment(vertical='top')
pyxl_alignment_wrap = Alignment(wrap_text=True)
pyxl_font_bold = Font(bold=True)
pyxl_font_red = Font(color="A00000",bold=True,size = "13")
pyxl_font_grn = Font(color="00A000",bold=True,size = "13")
pyxl_font_blu = Font(color="0000A0",bold=True,size = "13")
pyxl_font_orn = Font(color="FF6600",bold=True,size = "13")
pyxl_fill_green = PatternFill(start_color="E0FFF0", end_color="E0FFF0", fill_type = "solid")
# Warning: the form "PatternFill(bgColor="xxxxxx", fill_type = "solid")" returns black cells
pyxl_backcolor_red = PatternFill(start_color='FCCDBA', end_color='FCCDBA', fill_type = "solid")
pyxl_backcolor_orn = PatternFill(start_color='FBEAAB', end_color='FBEAAB', fill_type = "solid")
pyxl_backcolor_yel = PatternFill(start_color='FCFDC7', end_color='FCFDC7', fill_type = "solid")
pyxl_backcolor_blu = PatternFill(start_color='C5E2FF', end_color='C5E2FF', fill_type = "solid")
pyxl_cve_fills = [pyxl_backcolor_red,pyxl_backcolor_orn,pyxl_backcolor_yel,pyxl_backcolor_blu,None]

def pyxl_write_cell(ws,row_num,column_num,value,border=None,font=None,fill=None,alignment=None):
    cell = ws.cell(row=row_num, column=column_num)
    try:
        cell.value = value
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if font:
            cell.font = font
    except Exception as e:
        print("ERROR:(%d,%d):%s" % (row_num,column_num,e))
    # Optional next column return value
    return(column_num+1)

#################################
# do_page_report
#

def dict_get_value(dict,name,default):
    return dict[name] if name in dict else default

def short_name(name):
        return(name[name.index('/')+1:])

def do_page_report(audit_id, options, user_id):
    conn = SQL_CONNECT()
    cur = SQL_CURSOR(conn)

    #_log_args("DO_PACKAGE_VERSIONS_REPORT", hb_audit.name, options)
    options = {'format': 'xlsx', 'report_page': 'HarborCVETable', 'orderby': 'package', 'filter': 'is_severity:critical_not_base', 'search': 'CVE-2021-44228', 'default_orderby': 'name', 'filter_value': 'on'}

    #'orderby': '-harborartifact__harborrepository__name',

    records = dict_get_value(options,'records','')
    format = dict_get_value(options,'format', '')
    title = dict_get_value(options,'title', '')
    report_type = dict_get_value(options,'report_type', '')
    record_list = dict_get_value(options,'record_list', '')
    full_description = ('1' == dict_get_value(options,'full_description', '0'))

    delimiter = ','
    report_full_path = ''

    report_page = dict_get_value(options,'report_page', '')
    search = dict_get_value(options,'search', '')
    filter = dict_get_value(options,'filter', '')
    filter_value = dict_get_value(options,'filter_value', '')
    orderby = dict_get_value(options,'orderby', '')
    default_orderby = dict_get_value(options,'default_orderby', '')

    #
    scope = dict_get_value(options,'scope', '')

    report_path = '.'
    report_name = 'test_hb_cve_report.xlsx'
    report_full_path = os.path.join(SRT_REPORT_DIR,report_path,report_name)
    wb = Workbook()
    ws = wb.active
    ws.title = "CVE Summary"

    if not default_orderby:
        default_orderby = "name"

    # Assemble the query set
    selectedAudit = audit_id
    sql = f"""SELECT hbv.id as id,hbv.name as name,hbv.package as package,hbv.version as version,hbv.fix_version as fix_version,hbv.severity as severity,hbv.description as description,hbv.results_class as results_class,hbv.results_target as results_target, hbv.results_type as results_type, hbv.package_level as package_level, hr.repo_team,hr.repo_type, ocve.cvssV3_baseScore,ocve.cvssV2_baseScore,hbv.harborartifact_id,ha.tag,hr.name,hp.name FROM wr_studio_harborvulnerability as hbv  INNER JOIN wr_studio_harborartifact as ha ON hbv.harborartifact_id = ha.id INNER JOIN orm_cve as ocve ON hbv.cve_id = ocve.id INNER JOIN wr_studio_harborrepository hr ON ha.harborrepository_id = hr.id INNER JOIN wr_studio_harborproject hp ON hr.harborproject_id=hp.id WHERE  hbv.harboraudit_id is ?"""
    queryset = SQL_EXECUTE(cur, sql, (audit_id,)).fetchall()

    if search:
        pass
            # queries = None
            # search_term = search.replace('%20',' ').replace('%22','"').replace('%27','"').replace('+',' ')
            # for st in shlex.split(search_term):
            #     for field in HarborVulnerability.search_allowed_fields:
            #         if st.startswith('-'):
            #             query = ~Q(**{field + '__icontains': st[1:]})
            #         else:
            #             query = Q(**{field + '__icontains': st})
            #         # Debug support
            #         if False:
            #             my_query = queryset.filter(query)
            #             _log("FOOBAR:%s:%s" % (field,len(my_query)) )
            #         if queries:
            #             queries |= query
            #         else:
            #             queries = query
            # queryset = queryset.filter(queries)

            # Filters
    # if filter == "is_severity:Critical":
    #     queryset = queryset.filter(severity="Critical")
    # elif filter == "is_severity:critical_base":
    #     queryset = queryset.filter(Q(severity="Critical")&Q(package_level="1"))
    # elif filter == "is_severity:critical_not_base":
    #     queryset = queryset.filter(Q(severity="Critical")&~Q(package_level="1"))
    # elif filter == "is_severity:High":
    #     queryset = queryset.filter(severity="High")
    # elif filter == "is_severity:critical_high":
    #     queryset = queryset.filter(Q(severity="Critical")|Q(severity="High"))
    # elif filter == "is_severity:Medium":
    #     queryset = queryset.filter(severity="Medium")
    # elif filter_value == "is_severity:Low":
    #     queryset = queryset.filter(severity="Low")
    # elif filter == "is_severity:medium_low":
    #     queryset = queryset.filter(Q(severity="Medium")|Q(severity="Low"))
    # elif filter == "is_layer:layer1":
    #     queryset = queryset.filter(Q(package_level="1"))
    # elif filter == "is_layer:notlayer1":
    #     queryset = queryset.filter(~Q(package_level="1"))

    #queryset = HarborVulnerability.objects.filter(harboraudit=hb_audit)

    row = 1
    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(9)].width = 40

    # Header
    header = []
    header.extend([
        'Name',
        'Severity',
        'Package',
        'Version',
        'Fix Version',
        'Level',
        'Score',
        'Score V3',
        'Score V2',
        'Description',
        'Results Target',
        'Results Class',
        'Results Type',
        'Tag',
        'Repository',
        'Type',
        'Team',
        'Project',
    ])
    col = 1
    for name in header:
        pyxl_write_cell(ws,row,col,name,border=pyxl_border_bottom)
        col += 1
    row += 1

    parent_cache = {}
    progress_set_max(len(queryset))
    count = 0
    for id,hbv_name,package,version,fix_version,severity,description,results_class,results_target, results_type, package_level, repo_team,repo_type,cvssV3_baseScore,cvssV2_baseScore,harborartifact_id,tag,hr_name,hp_name in queryset:
        count += 1
        if 0 == (count % 1024):
            progress_set_current(count)
            progress_show(hbv_name,0)

        #print(hbv_name)
        # Use the cache for speedharborartifact
        if harborartifact_id in parent_cache:
            parent_data = parent_cache[harborartifact_id]
        else:
            parent_data = [tag,short_name(hr_name),repo_type,repo_team,hp_name]

            parent_cache[harborartifact_id] = parent_data
        # Generate the row
        vul_row = []

        if cvssV3_baseScore:
            score = cvssV3_baseScore if cvssV3_baseScore else cvssV2_baseScore
            score_v3 = cvssV3_baseScore
            score_v2 = cvssV2_baseScore
        else:
            score = ''
            score_v3 = ''
            score_v2 = ''

        vul_row.extend([
            hbv_name,
            severity,
            package,
            version,
            fix_version if fix_version else ' ',
            package_level if package_level else ' ',
            score,
            score_v3,
            score_v2,
            description if full_description else description[:30],
            results_target,
            results_class,
            results_type,
            parent_data[0],
            parent_data[1],
            parent_data[2],
            parent_data[3],
            parent_data[4],
        ])
        col = 1
        for hbv_name in vul_row:
            pyxl_write_cell(ws,row,col,hbv_name)
            col += 1
        # Color the severity
        if severity[0] == 'C':
            ws.cell(row=row,column=2).fill=pyxl_cve_fills[0]
        elif severity[0] == 'H':
            ws.cell(row=row,column=2).fill=pyxl_cve_fills[1]
        elif severity[0] == 'M':
            ws.cell(row=row,column=2).fill=pyxl_cve_fills[2]
        else:
            ws.cell(row=row,column=2).fill=pyxl_cve_fills[3]
        # Next row
        #break
        row += 1

    progress_done('Done')
    if report_full_path:
        wb.save(report_full_path)
    _log("DO_PAGE_REPORT:SCOPE=%s:REPORT=%s:FULL_DESC=%s:" % (scope,report_page,full_description))
    return report_full_path


#################################
# repo match test
# Purpose: find the matching audit artifact betweem two audits
#   This is to facilitate smart audit diffs, by matching up
#   audit artifacts that may have their versions updated
# Challenges:
#   In thoery, an audit would be a set of unique repos, each
#   with a specific artifact (usually selected by "tag"), and
#   in betweem audits only the version might change
#   1. However, in pratice, there are in many cases several tags
#   selected for a give repo, and matching those up can be
#   difficult, especially if there are also many tag updates.
#   2. Additionally, repos like DFL and OTA use tags to identify
#   sub-repos, where the tags are a combination of the sub-repo name
#   and the version.
#
# Resolution Model:
#   1. Collect all of the repo+tag pairs in each of the audits
#   2. Collect all of the cases where the repo+tag pairs are identical,
#      and this have not changed between audits
#   3. Iterate each remaining repo+tag pairs, and match them in order.
#      This is helped by presorting the lists in reverse tag string sort
#      order, so that the likelyhood a meanful matches is high, since
#      version string are usually in sort order format. In theory
#      the "push_time" could be used to order the potential matches, but
#      direct Trivy scans do not have this Harbor data, and the match
#      correctness would probably still not be any better.
#      In practice, most multiple tags are in pairs where most already
#      have one exact match, so this algorithm is reasonably as good as
#      possible without explicit manual intervention.
#

def scan_audit(audit_id,cur):
    sql = """SELECT * FROM wr_studio_harboraudit WHERE id = ?"""
    hb_audit = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()
    if hb_audit is None:
        print(f"ERROR: not such audit id '{audit_id}'")
        exit(1)

    artifact_table = []
    sql = """SELECT * FROM wr_studio_harborartifact WHERE harboraudit_id = ?"""
    for hb_art in SQL_EXECUTE(cur, sql, (audit_id,)).fetchall():

        # Ignore if this is not an active artifact
        if (not hb_art['selected_top']) and (1 != int(hb_art['audit_status'])):
            continue

        sql = """SELECT * FROM wr_studio_harborrepository WHERE id = ? ORDER BY '-tag';"""
        hb_repo = SQL_EXECUTE(cur, sql, (hb_art['harborrepository_id'],)).fetchone()

        repo_name = hb_repo['name']
        tag = hb_art['tag']
        push_time = hb_art['push_time']
        if not push_time: push_time = ''

        # Case (a) and (b): Key for normal single tag, and/or DFL tag
        key1 = repo_name
        if ('/dfl/' in repo_name) or ('/ota/' in repo_name):
#            print(f"FOO1:{repo_name}:{tag}")
#            print(f"   ARTID={hb_art['id']},NAME={hb_art['name']},Tags={hb_art['tag_names']}")
            key1 = f"{repo_name}|{tag[tag.find('-')]}"
        # Case (c): Obvious key for full repo+tag
        key2 = f"{repo_name}|{tag}"

        artifact_table.append([hb_art['id'],key1,key2,False])
    return(artifact_table)

# Audit_1 is the before audit
# Audit_2 is the current audit
# Return: (map_repos,added_repos,removed_repos)
#   map_repos: dictionary of current audit_artifact_ids to previous audit_artifact_ids
#   added_repos: list of added audit_artifact_ids (new to current release)
#   removed_repos: list of removed audit_artifact_ids (not found in current release)
def audits_repo_match(audit_1_id,audit_2_id,passed_conn=None):
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    # Audit 1
    artifact_table_1 = scan_audit(audit_1_id,cur)

    # Audit 2
    artifact_table_2 = scan_audit(audit_2_id,cur)

    # Audit matching
    map_table = []
    map_repos = {}
    added_repos = []
    removed_repos = []
    multiple_test = {}

    match_full = 0
    match_one  = 0
    match_next = 0
    match_none = 0
    match_remain = 0
    TAKEN = 3

    # First pass for absolute matches
    for i in range(len(artifact_table_2)):
        B_hb_art_id,B_key1,B_key2,B_taken = artifact_table_2[i]
        for j in range(len(artifact_table_1)):
            A_hb_art_id,A_key1,A_key2,A_taken = artifact_table_1[j]
            if B_key2 == A_key2:
                map_table.append([A_hb_art_id,B_hb_art_id,A_key2,B_key2,'0_Full tag match'])
                map_repos[B_hb_art_id] = A_hb_art_id
                artifact_table_2[i][TAKEN] = True
                artifact_table_1[j][TAKEN] = True
                match_full += 1
                break

    # Second pass for ordered repo name matches (not tags)
    for i in range(len(artifact_table_2)):
        B_hb_art_id,B_key1,B_key2,B_taken = artifact_table_2[i]
        if B_taken:
            continue
        for j in range(len(artifact_table_1)):
            A_hb_art_id,A_key1,A_key2,A_taken = artifact_table_1[j]
            if A_taken:
                continue
            if B_key1 == A_key1:
                if A_key1 in multiple_test:
                    reason = "2_Collision Next"
                    match_next += 1
                else:
                    reason = "1_Repo name match"
                    multiple_test[A_key1] = 1
                    match_one += 1
                map_table.append([A_hb_art_id,B_hb_art_id,A_key2,B_key2,reason])
                map_repos[B_hb_art_id] = A_hb_art_id
                artifact_table_2[i][TAKEN] = True
                artifact_table_1[j][TAKEN] = True
                break

    # Third pass for not found in A
    for i in range(len(artifact_table_2)):
        B_hb_art_id,B_key1,B_key2,B_taken = artifact_table_2[i]
        if B_taken:
            continue
        map_table.append([0,B_hb_art_id,'',B_key2,'3_Not Found'])
        added_repos.append(B_hb_art_id)
        match_none += 1

    # Fourth pass for remaing in A
    for i in range(len(artifact_table_1)):
        A_hb_art_id,A_key1,A_key2,A_taken = artifact_table_1[i]
        if A_taken:
            continue
        map_table.append([A_hb_art_id,0,A_key2,'','4_Remaining_in_A'])
        removed_repos.append(A_hb_art_id)
        match_remain += 1

    # Write results
    if verbose:
        print("ART_ID_PREV,ART_ID_NOW,KEY_PREV,KEY_NOW,REASON")
        for A_hb_art_id,B_hb_art_id,A_key2,B_key2,reason in map_table:
            print(f"{A_hb_art_id},{B_hb_art_id},{A_key2},{B_key2},{reason}")
        print(f"match_full={match_full}")
        print(f"match_one ={match_one}")
        print(f"match_next={match_next}")
        print(f"match_none={match_none}")
        print(f"match_remain={match_remain}")

    SQL_CLOSE_CUR(cur)
    if not passed_conn:
        SQL_CLOSE_CONN(conn)
    return(map_repos,added_repos,removed_repos)

# display_repo_match("<from_audit>,<to_audit>"), for example "521,568"
def display_repo_match(audits):
    audit_1_id,audit_2_id = audits.split(',')
    audits_repo_match(audit_1_id,audit_2_id)


#This fuction is created to populate CveOfTheDay model data
def fetch_cves_of_the_day(id_check,passed_conn=None):
    print('FETCH_CVES_OF_THE_DAY: called with audit id---->',id_check)
    audit_id_check = id_check
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)
    sql = f"""SELECT create_time,content FROM wr_studio_harboraudit WHERE id = ?;"""
    check_audit = SQL_EXECUTE(cur, sql,(audit_id_check,))
    check_list = []
    dateOfSelectedAuditstr = ''
    contentOfSelectedAudit = ''
    for row1 in check_audit:
        check_list.append(row1)
        dateOfSelectedAuditstr = (row1['create_time']).split()[0]
        contentOfSelectedAudit = row1['content']

    dateOfSelectedAudit = (datetime.strptime(dateOfSelectedAuditstr, '%Y-%m-%d')).date()
    print(check_list)
#    print('selected audit id', audit_id_check)
#    print('date of selected audit',dateOfSelectedAudit)
#    print('content of selected audit',contentOfSelectedAudit)
    dateOfpreviousDayAudit = (dateOfSelectedAudit-timedelta(days = 1))
    dateOf6DysBackAudit = (dateOfSelectedAudit-timedelta(days = 6))
#    print('date of previous day audit', dateOfpreviousDayAudit)
#    print('date of 6 days back audit', dateOf6DysBackAudit)

    sql1 = f"""SELECT id FROM wr_studio_harboraudit WHERE DATE(create_time) between ? AND ? AND content = ? ORDER BY id DESC ;"""
    auditsOfCurrentWeeksql = SQL_EXECUTE(cur, sql1,(dateOf6DysBackAudit,dateOfpreviousDayAudit,contentOfSelectedAudit))

    auditsOfCurrentWeek = []
    for auditID in auditsOfCurrentWeeksql:
        auditsOfCurrentWeek.append(auditID['id'])
    print('audits of current week',auditsOfCurrentWeek)

    #this function takes map_repos,added_repos,removed_repos returned by audits_repo_match() function as a parameter
    def (map_repos,added_repos,removed_repos):
        if False and verbose:
            print(map_repos)
            print(added_repos)
            print(removed_repos)

        new_cves_ids = []
        resolve_cves_ids = []

        # map_repos: dictionary of current audit_artifact_ids to previous audit_artifact_ids
        for k,v in map_repos.items():
            new_audit_cves = {}
            old_audit_cves = {}
            sql2 = """select id,name,package from wr_studio_harborvulnerability where harborartifact_id = ?"""
            new_cves = SQL_EXECUTE(cur, sql2,(k,)).fetchall()

            for i in new_cves:
                new_audit_cves[i['id']] = (i['name'],i['package'])

            old_cves = SQL_EXECUTE(cur, sql2,(v,)).fetchall()
            for j in old_cves:
                old_audit_cves[j['id']] = (j['name'],j['package'])
            #new cves check
            for k1,v1 in new_audit_cves.items():
                if v1 in old_audit_cves.values():
                    continue
                else:
                    new_cves_ids.append(k1)
            #resolve cves check
            for k2,v2 in old_audit_cves.items():
                if v2 in new_audit_cves.values():
                    continue
                else:
                    resolve_cves_ids.append(k2)

        # added_repos is a list of artifact IDs new in the current audit
        #Iterate added_repos, look up the respective vulnerability sets, and mark these CVEs as new
        for art in added_repos:
            sql = """select id from wr_studio_harborvulnerability where harborartifact_id = ?"""
            added_repos_cves = SQL_EXECUTE(cur, sql,(art,)).fetchall()

            for cve in added_repos_cves:
                new_cves_ids.append(cve['id'])

        # removed_repos is a list of artifact IDs removed since the previous audit
        # Iterate removed_repos, look up the respective vulnerability sets, and mark these CVEs as resolved (removed)
        for art1 in removed_repos:
            sql = """select id from wr_studio_harborvulnerability where harborartifact_id = ?"""
            removed_repos_cves = SQL_EXECUTE(cur, sql,(art1,)).fetchall()

            for cve in removed_repos_cves:
                resolve_cves_ids.append(cve['id'])

        return (new_cves_ids,resolve_cves_ids)

    new_cves_of_the_day = []
    resolve_of_the_day = []

    #new cves and removed cves of the day
    if len(auditsOfCurrentWeek) > 0:
        previousDayAuditID = auditsOfCurrentWeek[0]
        print('previous day audit id',previousDayAuditID)
        #calling audits_repo_match(audit_prev_id, audit_now_id)
        map_repos,added_repos,removed_repos = audits_repo_match(previousDayAuditID,audit_id_check,passed_conn=conn)
        new_cves_of_the_day1,resolve_of_the_day1 = (map_repos,added_repos,removed_repos)

        #"is_added" field of cveoftheday is True for new CVEs, or False for resolved CVEs
        for i in new_cves_of_the_day1:
            new_cves_of_the_day.append((i,audit_id_check,contentOfSelectedAudit,True))

        for j in resolve_of_the_day1:
            resolve_of_the_day.append((j,audit_id_check,contentOfSelectedAudit,False))

    print('  cves of the day--->',len(new_cves_of_the_day))
    print('  resolve of the day--->',len(resolve_of_the_day))

    sql4 = f""" select id from wr_studio_cveoftheday where harboraudit_id = ?;"""
    existrec = SQL_EXECUTE(cur, sql4,(audit_id_check,)).fetchall()
    print(len(existrec))

    if len(existrec ) == 0:
        if len(new_cves_of_the_day) > 0:
            SQL_BATCH_WRITE(cur,
                    "wr_studio_cveoftheday",
                    new_cves_of_the_day,
                    fields=["harbor_cve_id","harboraudit_id","content","is_added"],
                    override_values=None,
                    )
        if len(resolve_of_the_day) > 0:
            SQL_BATCH_WRITE(cur,
                    "wr_studio_cveoftheday",
                    resolve_of_the_day,
                    fields=["harbor_cve_id","harboraudit_id","content","is_added"],
                    override_values=None,
                    )
        SQL_COMMIT(conn)
        print('data inserted for audit_id',audit_id_check)

    if not passed_conn:
        SQL_CLOSE_CONN(conn)


#################################
# api_cve_table
#
# Generate a CVS file of the CVE status of a given audit
# Most data comes from the Trivy scans (CVE, package, version)
# The V3 and V3 scoring comes from NVD via the SRTool for WR Linux database
# Return the file zipped, to reduce the transfer size by 80%
#

def api_cve_table(parameters):
    audit_id,report_name = parameters.split(',')

    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    AUDIT_ON = 1
    TRUE = 1

    # Get the audit record
    sql = f"""SELECT * FROM wr_studio_harboraudit WHERE id = ?"""
    hb_audit = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()

    # Get the artifact list of record
    hb_art_list = []
    sql = f"""SELECT * FROM wr_studio_harborartifact WHERE harboraudit_id = ? AND audit_status = ?"""
    for hb_art in SQL_EXECUTE(cur, sql, (audit_id,AUDIT_ON,)).fetchall():
        hb_art_list.append(hb_art)
    sql = f"""SELECT * FROM wr_studio_harborartifact WHERE harboraudit_id = ? AND selected_top = ?"""
    for hb_art in SQL_EXECUTE(cur, sql, (audit_id,TRUE,)).fetchall():
        hb_art_list.append(hb_art)

    vul_table = []
    cve_cache = {}
    cve_missing = {}
    cve_missing_studio_count = 0
    cve_missing_wrlinx_count = 0
    for hb_art in hb_art_list:
        # Get the parent repository
        harborrepository_id = hb_art['harborrepository_id']
        sql = f"""SELECT * FROM wr_studio_harborrepository WHERE id = ?"""
        hb_repo = SQL_EXECUTE(cur, sql, (harborrepository_id,)).fetchone()

        # Get the parent project
        project_id = hb_repo['harborproject_id']
        sql = f"""SELECT * FROM wr_studio_harborproject WHERE id = ?"""
        hb_project = SQL_EXECUTE(cur, sql, (project_id,)).fetchone()

        # Get this artifact's vulnerabilities
        sql = f"""SELECT * FROM wr_studio_HarborVulnerability WHERE harborartifact_id = ?"""
        vulnerabilities = SQL_EXECUTE(cur, sql, (hb_art['id'],)).fetchall()
        for vul in vulnerabilities:
            cve_name = vul['name']
            # Load and cache CVE info
            if not cve_name in cve_cache:
                sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
                cve = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
                if cve:
                    score_v2 = cve['cvssV2_baseScore']
                    score_v3 = cve['cvssV3_baseScore']
                    score = score_v3 if score_v3 else score_v2
                else:
                    if cve_name.startswith('CVE'):
                        cve_missing[cve_name] = True
                        cve_missing_studio_count += 1
                        score_v2 = ''
                        score_v3 = ''
                    else:
                        score_v2 = ''
                        score_v3 = ''
                cve_cache[cve_name] = [score_v2,score_v3]
            score_v2,score_v3 = cve_cache[cve_name]

            # Append the vulnerability row
            vul_table.append([
                cve_name,
                vul['severity'],
                vul['package'],
                vul['version'],
                vul['fix_version'],
                vul['package_level'],

                score_v3,
                score_v2,

                vul['results_target'],
                vul['results_class'],
                vul['results_type'],

                hb_art['tag'],
                hb_repo['name'].replace(hb_project['name']+'/',''),
                hb_repo['base_image'],
                hb_repo['repo_type'],
                hb_repo['repo_team'],
                hb_project['name'],
            ])
    SQL_CLOSE_CONN(conn)

    #
    # Gather the missing CVE information
    #

    # Insert via refresh_cve_nist, to not waste time in this API call
    cve_writeback = {}
    if False:
        wrlx_dbconfig = {}
        wrlx_dbconfig['dbtype'] = "sqlite_wrlinux"
        wrlx_dbconfig['sqlite_wrlinux'] = {'path':'srt_wrlinux.sqlite'}
        print("FOO1")
        conn = SQL_CONNECT(column_names=True,srt_dbconfig=wrlx_dbconfig)
        cur = SQL_CURSOR(conn)
        print("FOO2")
        if not conn:
            print(f"ERROR:could not open db '{wrlx_dbconfig['sqlite_wrlinux']}'")
            exit(1)
        print(f"FOO3:{cve_missing_studio_count}")
        i = 0
        cve_writeback = {}
        for cve_name in cve_missing:
            if 0 == (i % 50):
                print(f"IN:{i:5}:{cve_name}")
            i += 1
            sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
            cve = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
            if cve:
                score_v2 = cve['cvssV2_baseScore']
                score_v3 = cve['cvssV3_baseScore']
                score = score_v3 if score_v3 else score_v2
                # Write back
                sql_values = (
                    cve['name'],
                    cve['name_sort'],
                    cve['priority'],
                    cve['status'],
                    cve['comments'],
                    cve['comments_private'],
                    cve['tags'],
                    cve['cve_data_type'],
                    cve['cve_data_format'],
                    cve['cve_data_version'],
                    cve['public'],
                    cve['publish_state'],
                    cve['publish_date'],
                    cve['acknowledge_date'],
                    cve['description'],
                    cve['publishedDate'],
                    cve['lastModifiedDate'],
                    cve['recommend'],
                    cve['recommend_list'],
                    cve['cvssV3_baseScore'],
                    cve['cvssV3_baseSeverity'],
                    cve['cvssV2_baseScore'],
                    cve['cvssV2_severity'],
                    cve['packages'],
                    cve['score_date'],
                    cve['srt_updated'],
                    cve['srt_created'],
                )
                cve_writeback[cve_name] = sql_values
            else:
                score = ' -'
                score_v2 = ' -'
                score_v3 = ' -'
                cve_missing_wrlinx_count += 1
            cve_cache[cve_name] = [score_v2,score_v3]
        SQL_CLOSE_CONN(conn)
        print("FOO4")

        #
        # Write back new CVE scores
        #

        conn = SQL_CONNECT(column_names=True)
        cur = SQL_CURSOR(conn)
        i = 0
        count_create = 0
        for cve_name in cve_writeback:
            if 0 == (i % 50):
                print(f"UP:{i:5}:{cve_name}")
                SQL_COMMIT(conn)
            i += 1

            sql = '''SELECT * FROM orm_cve WHERE name=?'''
            cve_current = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
            cve_id = -1
            srtool_today = datetime.today()
            if cve_current is None:
                # Insert via refresh_cve_nist, to not waste time in this API call
                pass
            else:
                cvssV2_baseScore,cvssV3_baseScore= cve_cache[cve_name]
                sql = ''' UPDATE orm_cve SET cvssV2_baseScore=?, cvssV3_baseScore=? WHERE name=?'''
                SQL_EXECUTE(cur, sql, (cvssV2_baseScore,cvssV3_baseScore,cve_name,))
        SQL_COMMIT(conn)
        SQL_CLOSE_CONN(conn)
        print(f"count_create={count_create}")

    #
    # Generate the report
    #

#    report_name = f"reports/vul_report_{hb_audit['content'].replace(' ','-')}_{hb_audit['create_time'][:10].replace('/','-')}.csv"
    report_full_name = f"reports/{report_name}.csv"
    report_zip_name = f"reports/{report_name}.zip"
    with open(report_full_name, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, delimiter=',',
                            quotechar='"', quoting=csv.QUOTE_MINIMAL)

        # Write the header
        writer.writerow([
            "Name",
            "Severity",
            "Package",
            "Version",
            "Fix Version",
            "Level",
            "Score V3",
            "Score V2",
            "Description",
            "Results Target",
            "Results Class",
            "Results Type",
            "Tag",
            "Repository",
            "Base Image",
            "Type",
            "Team",
            "Project",
        ])

        for vul in vul_table:
            # Override scores
            vul[7],vul[6]= cve_cache[vul[0]]
            # Write the vulnerability row
            writer.writerow(vul)

    # Compress it
    os.system(f"zip --junk-paths {report_zip_name} {report_full_name}")

    print(f"cve_missing_studio_count = {cve_missing_studio_count}")
#    print(f"cve_missing_wrlinx_count = {cve_missing_wrlinx_count}")
    print(f"Report file = {report_full_name}")
    print(f"ZIP    file = {report_zip_name}")

#################################
# refresh_cve_nist
#
# Update the active CVEs in the SRTool for Studio database.
# Draw the data from the SRTool for WR Linux database, which
# daily downloads the NVD updates, and includes the latest Jira
# information.

# Copy in the latest SRTool for WR Linux database
# Copy from the backed-ups to not disturb the running database
def import_srtool_db(local_srtool_db_path):
    returncode = 0

    if not srt_lx_path:
        print("ERROR: missing 'srt_lx_path'")
        exit(1)

    # Use "bash" to expand the wild card
    print(f"Find latest backed-up SRTool database at: {srt_lx_remote},{srt_lx_path}")
    cmd = ['bash','-c',f"ls -t {srt_lx_path}/*/srt.sqlite"]
    if srt_lx_remote:
        # "david@147.11.209.83"
        cmd = ['ssh',srt_lx_remote] + cmd
    returncode,stdout,stderr = execute_process(cmd)
    if returncode:
        print("%s:%s:%s" % (returncode,cmd,stderr))
        exit(1)
    srtool_database = stdout.split('\n')[0]
    print(f"Copy in SRTool database: {srtool_database}")
    if srt_lx_remote:
        cmnd = f"scp -p {srt_lx_remote}:{srtool_database} {local_srtool_db_path}"
    else:
        cmnd = f"cp --preserve=timestamps {srtool_database} {local_srtool_db_path}"
    print(f"CMND={cmnd}")
    returncode = os.system(cmnd)
    print("Copy done (%d)" % returncode)
    return srtool_database

def refresh_cve_nist(skip_db_copy=False,skip_to_cve=''):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # Quick debug
    skip = False
    if skip_to_cve:
        skip = True
        skip_cve=skip_to_cve    # next: CVE-2015-5237 CVE-2005-2541

    log_file = open('cve_update.log', 'a')
    def show_status(msg):
        if msg.startswith('---'):
            msg = f'{msg:30} ({datetime.today().strftime("%Y-%m-%d %M:%S")})'
        elif msg:
            msg = f'{msg:30} ({datetime.today().strftime("%M:%S")})'
        log_file.write(f"{msg}\n")
        print(msg)
    show_status('')
    show_status('--------------------------')

    #
    # Find and import the late SRTOol for WR Linux database
    #

    # Copy in current database for SRTool for LX
    srtool_dest_db_path = SRTOOL_DEST_DB_PATH
    if not skip_db_copy:
        srtool_source_db_file = import_srtool_db(srtool_dest_db_path)
    else:
        srtool_source_db_file = srtool_dest_db_path
    file_mod_time = time.ctime(os.path.getmtime(srtool_dest_db_path))
    show_status(f"SRTOOL_DEST_DB_FILE  ({file_mod_time})={srtool_source_db_file}")

    #
    # Scan latest audit types
    #

    cve_harbor_list = {}
    audit_list = []
    for content in ('conductor','conductor_latest','wr_studio_product_latest'):
        sql = f"""SELECT * FROM wr_studio_harboraudit WHERE content = ? ORDER BY id DESC ;"""
        audit = SQL_EXECUTE(cur, sql, (content,)).fetchone()
        if not audit:
            print("ERROR:No recent audit found for '{content}'")
            exit(1)
        audit_id = audit['id']
        show_status(f"AUDIT_ID = {audit_id} ({content})")
        audit_list.append(audit_id)

        # Gather the active CVE names of this audit type
        i = 0
        sql = f"""SELECT * FROM wr_studio_HarborVulnerability WHERE harboraudit_id = ?"""
        vulnerabilities = SQL_EXECUTE(cur, sql, (audit_id,)).fetchall()
        show_status(f"GET:COUNT={len(vulnerabilities)}")
        for vul in vulnerabilities:
            cve_name = vul['name'].strip()
            if 0 == (i % 100):
                print(f"GET:{i:5}:{cve_name}    \r",end='')
            i += 1
            if cve_name.startswith('CVE'):
                cve_harbor_list[vul['name']] = True

    print('')
    SQL_CLOSE_CONN(conn)

    #
    # Gather the current CVE information from WR Linux SRTool
    #

    wrlx_dbconfig = {}
    wrlx_dbconfig['dbtype'] = "sqlite_wrlinux"
    wrlx_dbconfig['sqlite_wrlinux'] = {'path':'srt_wrlinux.sqlite'}
    conn = SQL_CONNECT(column_names=True,srt_dbconfig=wrlx_dbconfig)
    cur = SQL_CURSOR(conn)
    if not conn:
        print(f"ERROR:could not open db '{wrlx_dbconfig['sqlite_wrlinux']}'")
        exit(1)
    show_status(f"IN:COUNT={len(cve_harbor_list)}")
    cve_wrlinux_data = []
    cve_wrlinux_missing = []

    if True:
        for i,cve_name in enumerate(cve_harbor_list):
            if skip and (not skip_cve == cve_name):
                continue
            if 0 == (i % 50):
                print(f"IN:{i:5}:{cve_name}    \r",end='')
            sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
            cve = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
            if cve:
                cve_wrlinux_data.append(cve)
                if skip: print(f"FOUND:{cve_name}  ")
            else:
                if skip: print(f"MISSING:{cve_name}  ")
                cve_wrlinux_missing.append(cve_name)
    #            cve_wrlinux_data.append(cve)
                pass
    else:
        cve_list = []
        def check_cve_list(cve_list):
            cve_list_found = {}
            # Select * From Equipment Where ID IN (2, 3, 4, 7, 11, 34);
            select_str = ['"'+cve_name+'"' for cve_name in cve_list].join(',')
            print(f"SELECT_STR={select_str}")
            exit(1)
            sql = f"""SELECT * FROM orm_cve WHERE name IN ({select_str})"""
            cves = SQL_EXECUTE(cur, sql,).fetchall()
            for cve in cves:
                cve_wrlinux_data.append(cve)
                cve_list_found[cve['name']] = True
                if skip: print(f"FOUND:{cve_name}  ")
            for cve_name in cve_list:
                if not cve_name in cve_list_found:
                    if skip: print(f"MISSING:{cve_name}  ")
                    cve_wrlinux_missing.append(cve_name)
        #            cve_wrlinux_data.append(cve)

        for i,cve_name in enumerate(cve_harbor_list):
            if skip and (not 'CVE-2022-35252' == cve_name):
                continue
            if 0 == (i % 50):
                print(f"IN:{i:5}:{cve_name}    \r",end='')
            cve_list.append(cve_name)
            if 10 <= len(cve_list):
                check_cve_list(cve_list)
                cve_list = []
        if len(cve_list):
            check_cve_list(cve_list)

    SQL_CLOSE_CONN(conn)
    print('')

    #
    # Update Harbor CVEs with latest WR Linux data
    #

    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)
    cve_created_list = []
    cve_updated_list = []
    cve_noscore_list = {}
    cve_rec_map = {}
    show_status(f"UP:COUNT={len(cve_wrlinux_data)}")
    for i,cve in enumerate(cve_wrlinux_data):
        cve_name = cve['name']
        if 0 == (i % 50):
            print(f"UP:{i:5}:{cve_name}    \r",end='')
            if not test:
                SQL_COMMIT(conn)

        sql = '''SELECT * FROM orm_cve WHERE name=?'''
        cve_studio = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
        if cve_studio is None:
            cve_created_list.append(cve_name)
            sql_elements = [
                'name',
                'name_sort',
                'priority',
                'status',
                'comments',
                'comments_private',
                'tags',
                'cve_data_type',
                'cve_data_format',
                'cve_data_version',
                'public',
                'publish_state',
                'publish_date',
                'acknowledge_date',
                'description',
                'publishedDate',
                'lastModifiedDate',
                'recommend',
                'recommend_list',
                'cvssV3_baseScore',
                'cvssV3_baseSeverity',
                'cvssV2_baseScore',
                'cvssV2_severity',
                'packages',
                'score_date',
                'srt_updated',
                'srt_created',
                ]
            sql_qmarks = []
            for idx in range(len(sql_elements)):
                sql_qmarks.append('?')
            sql_values = [
                cve['name'],
                cve['name_sort'],
                cve['priority'],
                cve['status'],
                cve['comments'],
                cve['comments_private'],
                cve['tags'],
                cve['cve_data_type'],
                cve['cve_data_format'],
                cve['cve_data_version'],
                cve['public'],
                cve['publish_state'],
                cve['publish_date'],
                cve['acknowledge_date'],
                cve['description'],
                cve['publishedDate'],
                cve['lastModifiedDate'],
                cve['recommend'],
                cve['recommend_list'],
                cve['cvssV3_baseScore'],
                cve['cvssV3_baseSeverity'],
                cve['cvssV2_baseScore'],
                cve['cvssV2_severity'],
                cve['packages'],
                cve['score_date'],
                cve['srt_updated'],
                cve['srt_created'],
                ]
            if skip: print(f"FOO:ADD_CVE:{cve_name}:{cve['cvssV3_baseScore']}:{cve['cvssV3_baseSeverity']}:{cve['cvssV2_baseScore']}:{cve['cvssV2_severity']}")
            sql, params = 'INSERT INTO orm_cve (%s) VALUES (%s)' % (','.join(sql_elements),','.join(sql_qmarks)),sql_values
            if not test:
                SQL_EXECUTE(cur, sql, params)
                cve_rec_map[cve_name] = [SQL_GET_LAST_ROW_INSERTED_ID(cur),cve['publishedDate']]
            if (not cve['cvssV3_baseScore']) and (not cve['cvssV2_baseScore']):
                cve_noscore_list[cve_name] = "Yes/New WRL"
        else:
            is_cve_diff = \
                (cve['publishedDate']       != cve_studio['publishedDate']      ) or \
                (cve['lastModifiedDate']    != cve_studio['lastModifiedDate']   ) or \
                (cve['packages']            != cve_studio['packages']           )
            # If WR Linux db missing values, do not overwrite any WR Studio values
            if cve['cvssV3_baseScore'] or cve['cvssV2_baseScore']:
                is_cve_diff |= \
                    (cve['cvssV3_baseScore']    != cve_studio['cvssV3_baseScore']   ) or \
                    (cve['cvssV3_baseSeverity'] != cve_studio['cvssV3_baseSeverity']) or \
                    (cve['cvssV2_baseScore']    != cve_studio['cvssV2_baseScore']   ) or \
                    (cve['cvssV2_severity']     != cve_studio['cvssV2_severity']    )

            if skip: print(f"FOO:IS_CVE_DIFF:LX={is_cve_diff}:{cve['cvssV3_baseScore']}:{cve['cvssV3_baseSeverity']}:{cve['cvssV2_baseScore']}:{cve['cvssV2_severity']}")
            if skip: print(f"FOO:IS_CVE_DIFF:ST={is_cve_diff}:{cve_studio['cvssV3_baseScore']}:{cve_studio['cvssV3_baseSeverity']}:{cve_studio['cvssV2_baseScore']}:{cve_studio['cvssV2_severity']}")

            if is_cve_diff:
                cve_updated_list.append(cve_name)
                sql = ''' UPDATE orm_cve SET
                    publishedDate=?,
                    lastModifiedDate=?,
                    cvssV3_baseScore=?,
                    cvssV3_baseSeverity=?,
                    cvssV2_baseScore=?,
                    cvssV2_severity=?,
                    packages=?
                    WHERE id=?'''
                if not test:
                    SQL_EXECUTE(cur, sql, (
                        cve['publishedDate']      ,
                        cve['lastModifiedDate']   ,
                        cve['cvssV3_baseScore']   ,
                        cve['cvssV3_baseSeverity'],
                        cve['cvssV2_baseScore']   ,
                        cve['cvssV2_severity']    ,
                        cve['packages']           ,
                        cve_studio['id'],))
                    cve_rec_map[cve_name] = [cve_studio['id'],cve['publishedDate']]

            if (not cve['cvssV3_baseScore']) and (not cve['cvssV2_baseScore']):
                cve_noscore_list[cve_name] = "Yes WRL"

    #
    # Check the remaining Harbor CVEs for no scores
    #

    for cve_name in cve_wrlinux_missing:
        sql = '''SELECT * FROM orm_cve WHERE name=?'''
        cve_studio = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
        if not cve_studio:
            cve_noscore_list[cve_name] = "Not WRL"
        elif (not cve_studio['cvssV3_baseScore']) and (not cve_studio['cvssV2_baseScore']):
            cve_noscore_list[cve_name] = "Yes Studio"

    #
    # Assert orm_cve links in vulnerabilities
    #

    vul_no_cve = 0
    vul_had_cve = 0
    vul_look_cve = 0
    vul_set_cve = 0
    vul_miss_cve = {}
    sla_list = {}
    for audit_id in audit_list:
        i = 0
        sql = f"""SELECT * FROM wr_studio_HarborVulnerability WHERE harboraudit_id = ? AND cve_id IS ?"""
        vulnerabilities = SQL_EXECUTE(cur, sql, (audit_id,None,)).fetchall()
        print(f"GET:COUNT={len(vulnerabilities)}")
        for vul in vulnerabilities:
            cve_name = vul['name'].strip()
            if not cve_name.startswith("CVE-"):
                continue
            vul_no_cve += 1

            if not cve_name in cve_rec_map:
                sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
                cve_obj = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
                if cve_obj:
                    cve_id = cve_obj['id']
                    cve_rec_map[cve_name] = [cve_id,cve_obj['publishedDate']]
                    vul_look_cve += 1
                else:
                    cve_id = 0
                    vul_miss_cve[cve_name] = True
            else:
                cve_id = cve_rec_map[cve_name][0]
                vul_had_cve += 1

            if cve_id and (not test):
                sql = ''' UPDATE wr_studio_HarborVulnerability SET cve_id=? WHERE id=?'''
                SQL_EXECUTE(cur, sql, (cve_id,vul['id'],))
                vul_set_cve += 1
                sla_list[vul['id']] = cve_rec_map[cve_name][0]

            if 0 == (i % 100):
                print(f"CVE_SYNC[{audit_id}]:{i:5}:{cve_name}    \r",end='')
                SQL_COMMIT(conn)
            i += 1
        print("")
    SQL_COMMIT(conn)

    #
    # Also update the SLA/Age records
    #

    sql = f"""SELECT * FROM wr_studio_cvesla WHERE cve_id IS ?"""
    cve_slas = SQL_EXECUTE(cur, sql, (None,)).fetchall()
    print(f"SLA:COUNT={len(cve_slas)}")
    sla_update = 0
    i = 0
    for sla in cve_slas:
        cve_name = sla['name'].strip()
        if not cve_name.startswith("CVE-"):
            continue

        if not cve_name in cve_rec_map:
            sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
            cve_obj = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
            if cve_obj:
                cve_id = cve_obj['id']
                publishedDate = cve_obj['publishedDate']
                cve_rec_map[cve_name] = [cve_id,publishedDate]
            else:
                cve_id,publishedDate = 0,''
        else:
            cve_id,publishedDate = cve_rec_map[cve_name]

        if cve_id and (not test):
            sql = ''' UPDATE wr_studio_cvesla SET cve_id=?, nvd_date=? WHERE id=?'''
            SQL_EXECUTE(cur, sql, (cve_id,publishedDate,sla['id'],))
            sla_update += 1

        if 0 == (i % 100):
            print(f"SLA_SYNC[{audit_id}]:{i:5}:{cve_name}    \r",end='')
            SQL_COMMIT(conn)
        i += 1
    print("")
    print(f"sla_update   = {sla_update}")
    SQL_COMMIT(conn)

    #
    # Summary
    #

    if not test:
        SQL_COMMIT(conn)
    SQL_CLOSE_CONN(conn)
    show_status(f"Count_CVEs          = {len(cve_harbor_list)}")
    cve_wrlinux_missing_count = len(cve_wrlinux_missing)
    cve_updated_list_count = len(cve_updated_list)
    cve_created_list_count = len(cve_created_list)
    cve_noscore_list_count = len(cve_noscore_list)
    show_status(f"cve_wrlinux_missing = {cve_wrlinux_missing_count}")
    show_status(f"cve_updated_list    = {cve_updated_list_count}")
    show_status(f"cve_created_list    = {cve_created_list_count}")
    show_status(f"cve_noscore_list    = {cve_noscore_list_count}")

    show_status("=== Added from WR Linux ===")
    for i,cve_name in enumerate(cve_created_list):
        show_status(f"  {cve_name}")
        if i > 10:
            show_status("  ...")
            break
    show_status("=== Updated from WR Linux ===")
    for i,cve_name in enumerate(cve_updated_list):
        show_status(f"  {cve_name}")
        if i > 10:
            show_status("  ...")
            break
    cve_missing_file = open('cve_wrlinux_missing.log', 'w')
    show_status("=== Not found in WR Linux ===")
    cve_missing_file.write("=== Not found in WR Linux ===\n")
    for i,cve_name in enumerate(cve_wrlinux_missing):
        cve_missing_file.write(f"{cve_name}\n")
        if i < 40:
            show_status(f"  {cve_name}")
        if i == 40:
            print("  ...")
    show_status("=== CVEs no scores ===")
    cve_missing_file.write("=== CVEs no scores ===\n")
    for i,cve_name in enumerate(cve_noscore_list):
        # Double check that NVD does not have score
        # TODO #### PARAMETERIZE
        year = cve_name.split('-')[1]
        cmnd = ['./bin/nist/srtool_nist.py','--cve-detail',cve_name,'--file',f"/home/david/srtool-support/data/nvdcve-1.1-{year}.json"]
        exec_returncode,exec_stdout,exec_stderr = execute_process(cmnd)
        msg = []
        for line in exec_stdout.split('\n'):
            if 'baseScore' in line:
                msg.append(line)
            if 'cvssV3_baseSeverity' in line:
                msg.append(line)
            if 'cvssV2_severity' in line:
                msg.append(line)
        result = f"  {cve_name} ({cve_noscore_list[cve_name]}) (NVD:{','.join(msg)})"
        cve_missing_file.write(f"{result}\n")
        if i < 40:
            show_status(result)
        if i == 40:
            print("  ...")
    print("Full missing CVE report at:'cve_wrlinux_missing.log'")


def cve_link_update(audit_ids):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    if 'default' == audit_ids:
        audit_list = []
        for content in ('conductor','conductor_latest','wr_studio_product_latest'):
            sql = f"""SELECT * FROM wr_studio_harboraudit WHERE content = ? ORDER BY id DESC ;"""
            audit = SQL_EXECUTE(cur, sql, (content,)).fetchone()
            if not audit:
                print("ERROR:No recent audit found for '{content}'")
                exit(1)
            audit_id = audit['id']
            print(f"AUDIT_ID = {audit_id} ({content})")
            audit_list.append(audit_id)
    else:
        audit_list = audit_ids.split(',')

    #
    # Assert orm_cve links in vulnerabilities
    #

    cve_rec_map = {}
    vul_no_cve = 0
    vul_had_cve = 0
    vul_look_cve = 0
    vul_set_cve = 0
    vul_miss_cve = {}
    sla_list = {}
    for audit_id in audit_list:
        i = 0
        sql = f"""SELECT * FROM wr_studio_HarborVulnerability WHERE harboraudit_id = ? AND cve_id IS ?"""
        vulnerabilities = SQL_EXECUTE(cur, sql, (audit_id,None,)).fetchall()
        print(f"GET:COUNT={len(vulnerabilities)}")
        for vul in vulnerabilities:
            cve_name = vul['name'].strip()
            if not cve_name.startswith("CVE-"):
                continue
            vul_no_cve += 1

            if not cve_name in cve_rec_map:
                sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
                cve_obj = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
                if cve_obj:
                    cve_id = cve_obj['id']
                    cve_rec_map[cve_name] = [cve_id,cve_obj['publishedDate']]
                    vul_look_cve += 1
                else:
                    cve_id = 0
                    vul_miss_cve[cve_name] = True
            else:
                cve_id = cve_rec_map[cve_name][0]
                vul_had_cve += 1

            if cve_id and (not test):
                sql = ''' UPDATE wr_studio_HarborVulnerability SET cve_id=? WHERE id=?'''
                SQL_EXECUTE(cur, sql, (cve_id,vul['id'],))
                vul_set_cve += 1
                sla_list[vul['id']] = cve_rec_map[cve_name]

            if 0 == (i % 100):
                print(f"CVE_SYNC[{audit_id}]:{i:5}:{cve_name}    \r",end='')
                SQL_COMMIT(conn)
            i += 1
        print("")
    SQL_COMMIT(conn)


    # Also update the SLA/Age records
    sql = f"""SELECT * FROM wr_studio_cvesla WHERE cve_id IS ?"""
    cve_slas = SQL_EXECUTE(cur, sql, (None,)).fetchall()
    print(f"SLA:COUNT={len(cve_slas)}")
    sla_update = 0
    i = 0
    for sla in cve_slas:
        cve_name = sla['name'].strip()
        if not cve_name.startswith("CVE-"):
            continue

        if not cve_name in cve_rec_map:
            sql = f"""SELECT * FROM orm_cve WHERE name = ?"""
            cve_obj = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
            if cve_obj:
                cve_id = cve_obj['id']
                publishedDate = cve_obj['publishedDate']
                cve_rec_map[cve_name] = [cve_id,publishedDate]
            else:
                cve_id,publishedDate = 0,''
        else:
            cve_id,publishedDate = cve_rec_map[cve_name]

        if cve_id and (not test):
            sql = ''' UPDATE wr_studio_cvesla SET cve_id=?, nvd_date=? WHERE id=?'''
            SQL_EXECUTE(cur, sql, (cve_id,publishedDate,sla['id'],))
            sla_update += 1

        if 0 == (i % 100):
            print(f"SLA_SYNC[{audit_id}]:{i:5}:{cve_name}    \r",end='')
            SQL_COMMIT(conn)
        i += 1
    print("")
    print(f"sla_update   = {sla_update}")
    SQL_COMMIT(conn)


    print("")
    print(f"vul_no_cve   = {vul_no_cve}")
    print(f"vul_had_cve  = {vul_had_cve}")
    print(f"vul_look_cve = {vul_look_cve}")
    print(f"vul_set_cve  = {vul_set_cve}")
    print(f"vul_miss_cve = {len(vul_miss_cve)}:({vul_miss_cve.keys()})")

#################################
# check_wrlinux_cve
#

def check_wrlinux_cve(cve_list):
    wrlx_dbconfig = {}
    wrlx_dbconfig['dbtype'] = "sqlite_wrlinux"
    wrlx_dbconfig['sqlite_wrlinux'] = {'path':'srt_wrlinux.sqlite'}
    conn = SQL_CONNECT(column_names=True,srt_dbconfig=wrlx_dbconfig)
    cur = SQL_CURSOR(conn)
    if not conn:
        print(f"ERROR:could not open db '{wrlx_dbconfig['sqlite_wrlinux']}'")
        exit(1)
    i = 0
    for cve_name in cve_list.split(','):
        sql = '''SELECT * FROM orm_cve WHERE name=?'''
        cve = SQL_EXECUTE(cur, sql, (cve_name,)).fetchone()
        if cve:
            print(f"FOUND:{cve_name}:{cve['cvssV3_baseScore']}:{cve['cvssV3_baseSeverity']}:{cve['cvssV2_baseScore']}:{cve['cvssV2_severity']}")
        else:
            print(f"MISS :{cve_name}")
    SQL_CLOSE_CONN(conn)


#################################
# populate_cod_historical_data
#

def call_fetch_cves_of_the_day(id1,id2,passed_conn=None):
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    # Fetch table of audit IDs to skip (if any)
    audit_skip = {}
    file_fd = open('data/cve_age_repo_map.csv', 'r')
    for line in file_fd.readlines():
        line = line.strip()
        if (not line) or ('#' == line[0]):
            continue
        action,a,b = line.split(',')
        if 'SKIP_COD' == action:
            audit_skip[int(a)] = True
    if verbose:
        print(f"SKIP={audit_skip.keys()}")

    sql1 = """SELECT id FROM wr_studio_harboraudit WHERE id BETWEEN ? AND ? ORDER BY id;"""
    audit_ids_sql = SQL_EXECUTE(cur, sql1,(id1,id2)).fetchall()
    audit_ids = []
    for aud in audit_ids_sql:
        audit_ids.append(aud['id'])

    for audit_id in audit_ids:
        if audit_id in audit_skip:
            print(f"* SKIP audit #{audit_id}")
            continue
    
        sql2 = f"""select id from wr_studio_cveoftheday where harboraudit_id = ?;"""
        exist_rec = SQL_EXECUTE(cur, sql2,(audit_id,)).fetchall()

        if len(exist_rec) == 0:
            current_first_cod_audit = int(srtsetting_get(conn,'HARBOR_COD_FIRST','99999'))
            current_last_cod_audit = int(srtsetting_get(conn,'HARBOR_COD_LAST','1'))

            fetch_cves_of_the_day(audit_id,passed_conn=conn)

            if audit_id < current_first_cod_audit:
                srtsetting_set(conn,'HARBOR_COD_FIRST',str(audit_id))
            if audit_id > current_last_cod_audit:
                srtsetting_set(conn,'HARBOR_COD_LAST',str(audit_id))

    if not passed_conn:
        SQL_CLOSE_CONN(conn)

def populate_cod_historical_data(passed_val):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    if 'default' == passed_val:
        current_last_cod_audit = int(srtsetting_get(conn,'HARBOR_COD_LAST','1'))
        # find latest audit id
        sql1 = """SELECT id,name FROM wr_studio_harboraudit ORDER BY id DESC;"""
        hb_audit_top = SQL_EXECUTE(cur, sql1,()).fetchone()
        latest_audit_id = hb_audit_top['id']

        if current_last_cod_audit == latest_audit_id:
            print('cveoftheday data is upto date')
        elif current_last_cod_audit < latest_audit_id:
            call_fetch_cves_of_the_day(current_last_cod_audit,latest_audit_id,passed_conn=conn)

    else:
        ranges = passed_val.split(',')
        current_last_cod_audit = int(ranges[0])
        latest_audit_id = int(ranges[1])
        call_fetch_cves_of_the_day(current_last_cod_audit,latest_audit_id,passed_conn=conn)

    SQL_CLOSE_CONN(conn)


# Delete all records from cveoftheday table
# ./bin/wr_studio/srtool_cves.py --delete-cves-of-the-day 521   # "BASELINE 22.09 11-17-2022"
def cve_cod_delete(audit_base_id):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # delete all records from wr_studio_cveoftheday table
    sql = 'DELETE FROM wr_studio_cveoftheday;'
    SQL_EXECUTE(cur, sql, )
    SQL_COMMIT(conn)
    print('data deleted from cveoftheday table')

    #resetting
    srtsetting_set(conn,'HARBOR_COD_FIRST',audit_base_id)
    srtsetting_set(conn,'HARBOR_COD_LAST','1')

    SQL_COMMIT(conn)
    SQL_CLOSE_CONN(conn)


#################################
# dump_cod_historical_data
#
# parameters: baseline_id,target_is,content_string[#skip_id_1[,skip_id_2]*]
#
# Command line tests:
# ./bin/wr_studio/srtool_cves.py --dump-cod-historical-data 521,725,conductor#690
# ./bin/wr_studio/srtool_cves.py --dump-cod-historical-data 522,723,conductor_latest
# ./bin/wr_studio/srtool_cves.py --dump-cod-historical-data 524,713,wr-studio-product_latest
#

def dump_cod_historical_data(data,skip_list=''):
    print(f"DUMP_COD_HISTORICAL_DATA({data},{skip_list}")
    audit_baseline,audit_now,audit_content = data.split(',')
    skip_id_dict = {}
    if skip_list:
        print(f"SKIP_LIST:{skip_list}")
        for audit_id in [int(id) for id in skip_list.split(',')]:
            skip_id_dict[audit_id] = True
            print(f"SKIP:{audit_id}")

    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # Gather all of the audit IDs captured in the CoD table
    cod_found_audits = {}
    sql = f"""select harboraudit_id from wr_studio_cveoftheday;"""
    for hb_audit in SQL_EXECUTE(cur, sql,).fetchall():
        cod_found_audits[int(hb_audit['harboraudit_id'])] = True
    if verbose: print(f"* Found Harbor ID count={len(cod_found_audits)}")

    # Print the header
    print("CoR : Added CVEs for the audit, RoR = Resolved of audit, delta = CoR-RoR")
    print("Audit   Total  Running     CoR     RoR Date       Artifacts Content")
    print("------------------------------------------------------------------------")

    # Print the baseline summary
    sql = f"""SELECT * FROM wr_studio_harboraudit WHERE id = ?"""
    hb_audit = SQL_EXECUTE(cur, sql, (audit_baseline,)).fetchone()
    date = hb_audit['create_time'][:10]
    content = hb_audit['content']
    sql = f"""SELECT count(*) FROM wr_studio_harborvulnerability WHERE harboraudit_id = ?"""
    cve_total_baseline = SQL_EXECUTE(cur, sql, (audit_baseline,)).fetchone()[0]
    print(f"#{int(audit_baseline):4} {cve_total_baseline:7} {'':8} {'':7} {'':7} {date} {'':9} {content}")

    # Find and disable earlier duplicate date audits (last one wins), plus skip deleted audits
    date_last = ''
    audit_list = []
    date_skip_list = {}
    for audit_id in range(int(audit_now),int(audit_baseline),-1):
        sql = f"""SELECT create_time,content FROM wr_studio_harboraudit WHERE id = ?"""
        hb_audit = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()
        if not hb_audit:
            # This audit_id was deleted
            continue
        date = hb_audit['create_time'][:10]
        content = hb_audit['content']
        if content != audit_content:
            continue
        if date_last == date:
            date_skip_list[audit_id] = True
        date_last = date
        audit_list.append(audit_id)
    # Place the found audit list into accending order for cummulative CVE counts
    audit_list.sort()

    # Scan the found audits in accending order, compute running CVE totals
    running_total = cve_total_baseline
    for audit_id in audit_list:
        status = ''

        if not audit_id in cod_found_audits:
            print(f"#{audit_id:4} {0:7} {0:>8} {0:7} {0:7} {date} {0:9} {''} SKIPPED")
            continue

        # Get the audit details
        sql = f"""SELECT content,create_time FROM wr_studio_harboraudit WHERE id = ?"""
        hb_audit = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()
        if not hb_audit:
            print(f"ERROR:NO AUDIT:{audit_id}")
            continue
        content = hb_audit['content']
        date = hb_audit['create_time'][:10]

        if audit_id in skip_id_dict:
            status = "SKIP_AUDIT"
        elif audit_id in date_skip_list:
            status = "SKIP_OLDER_DUPLICATE_DATE"

        # Fetch Total number of CVEs of the audit
        sql = f"""SELECT count(*) FROM wr_studio_harborvulnerability WHERE harboraudit_id = ?"""
        cve_total = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()[0]

        # Fetch the total number of artifacts in this audit
        sql = f"""SELECT count(*) FROM wr_studio_harborartifact WHERE harboraudit_id = ? AND (selected_top = True OR audit_status = 1)"""
        artifact_total = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()[0]

        # Fetch CVEs of Product total
        sql = f"""SELECT count(*) FROM wr_studio_cveoftheday WHERE harboraudit_id = ? AND content = ? AND is_added = True"""
        cor_total = SQL_EXECUTE(cur, sql, (audit_id,audit_content,)).fetchone()[0]

        # Fetch resolved CVEs of Product total
        sql = f"""SELECT count(*) FROM wr_studio_cveoftheday WHERE harboraudit_id = ? AND content = ? AND is_added = False"""
        ror_total = SQL_EXECUTE(cur, sql, (audit_id,audit_content,)).fetchone()[0]

        # Compute the running total (if not skipped)
        new_running_total = running_total - ror_total + cor_total
        if not status:
            running_total = new_running_total
            if running_total != cve_total:
                status = f"(Offeset = {cve_total-running_total})"
        else:
            new_running_total = f"({running_total})"

        print(f"#{audit_id:4} {cve_total:7} {new_running_total:>8} {cor_total:7} {ror_total:7} {date} {artifact_total:9} {content} {status}")


#################################
# dump_cod_audit_data
#
# parameters: baseline_id,target_is,content_string[#skip_id_1[,skip_id_2]*]
#
# Command line tests:
# ./bin/wr_studio/srtool_cves.py --dump-cod-audit-data 592
#

def dump_cod_audit_data(audit_id):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # Added CVEs
    add_count = 0
    add_uniq_name = {}
    add_uniq_name_package = {}
    add_uniq_name_package_art = {}
    sql = f"""SELECT * FROM wr_studio_cveoftheday WHERE harboraudit_id = ? AND is_added = True"""
    for cod_add in SQL_EXECUTE(cur, sql, (audit_id,)).fetchall():
        add_count += 1
        sql = f"""SELECT * FROM wr_studio_harborvulnerability WHERE id = ?"""
        hb_vul = SQL_EXECUTE(cur, sql, (cod_add['harbor_cve_id'],)).fetchone()
        cve_name = hb_vul['name']
        package = hb_vul['package']
        art_id = hb_vul['harborartifact_id']
        add_uniq_name[cve_name] = True
        add_uniq_name_package[f"{cve_name}|{package}"] = True
        add_uniq_name_package_art[f"{cve_name}|{package}|{art_id}"] = True
        print(f"ADD:{cve_name:20},{package:30},{art_id:7}")

    # Resolved CVEs
    del_count = 0
    del_uniq_name = {}
    del_uniq_name_package = {}
    del_uniq_name_package_art = {}
    sql = f"""SELECT * FROM wr_studio_cveoftheday WHERE harboraudit_id = ? AND is_added = False"""
    for cod_del in SQL_EXECUTE(cur, sql, (audit_id,)).fetchall():
        del_count += 1
        sql = f"""SELECT * FROM wr_studio_harborvulnerability WHERE id = ?"""
        hb_vul = SQL_EXECUTE(cur, sql, (cod_del['harbor_cve_id'],)).fetchone()
        cve_name = hb_vul['name']
        package = hb_vul['package']
        art_id = hb_vul['harborartifact_id']
        del_uniq_name[cve_name] = True
        del_uniq_name_package[f"{cve_name}|{package}"] = True
        del_uniq_name_package_art[f"{cve_name}|{package}|{art_id}"] = True
        print(f"DEL:{cve_name:20},{package:30},{hb_vul['harborartifact_id']:7}")

    print(f"Count={add_count},Uniq_name={len(add_uniq_name)},Uniq_name_pkg={len(add_uniq_name_package)},Uniq_name_pkg_art={len(add_uniq_name_package_art)}")
    print(f"Count={del_count},Uniq_name={len(del_uniq_name)},Uniq_name_pkg={len(del_uniq_name_package)},Uniq_name_pkg_art={len(del_uniq_name_package_art)}")
    SQL_CLOSE_CONN(conn)

#################################
# review_cod_historical_data
#
# parameters: baseline_id,target_is,content_string[#skip_id_1[,skip_id_2]*]
#
# Command line tests:
# ./bin/wr_studio/srtool_cves.py --dump-cod-historical-data 521,725,conductor#690
# ./bin/wr_studio/srtool_cves.py --dump-cod-historical-data 522,723,conductor_latest
# ./bin/wr_studio/srtool_cves.py --dump-cod-historical-data 524,713,wr-studio-product_latest
#

def review_cod_historical_data(data,skip_list=''):
    print(f"DUMP_COD_HISTORICAL_DATA({data},{skip_list}")
    audit_baseline,audit_now,audit_content = data.split(',')
    skip_id_dict = {}
    if skip_list:
        print(f"SKIP_LIST:{skip_list}")
        for audit_id in [int(id) for id in skip_list.split(',')]:
            skip_id_dict[audit_id] = True
            print(f"SKIP:{audit_id}")

    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # Print the header
    print("Audit   Total  Running     CoR     RoR Date       Artifacts Content")
    print("------------------------------------------------------------------------")

    # Print the baseline summary
    sql = f"""SELECT * FROM wr_studio_harboraudit WHERE id = ?"""
    hb_audit = SQL_EXECUTE(cur, sql, (audit_baseline,)).fetchone()
    date = hb_audit['create_time'][:10]
    content = hb_audit['content']
    sql = f"""SELECT count(*) FROM wr_studio_harborvulnerability WHERE harboraudit_id = ?"""
    cve_total_baseline = SQL_EXECUTE(cur, sql, (audit_baseline,)).fetchone()[0]
    print(f"#{int(audit_baseline):4} {cve_total_baseline:7} {'':8} {'':7} {'':7} {date} {'':9} {content}")

    # Find and disable earlier duplicate date audits (last one wins), plus skip deleted audits
    date_last = ''
    audit_list = []
    date_skip_list = {}
    for audit_id in range(int(audit_now),int(audit_baseline),-1):
        sql = f"""SELECT create_time,content FROM wr_studio_harboraudit WHERE id = ?"""
        hb_audit = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()
        if not hb_audit:
            # This audit_id was deleted
            continue
        date = hb_audit['create_time'][:10]
        content = hb_audit['content']
        if content != audit_content:
            continue
        if date_last == date:
            date_skip_list[audit_id] = True
        date_last = date
        audit_list.append(audit_id)
    # Place the found audit list into accending order for cummulative CVE counts
    audit_list.sort()

    # Scan the found audits in accending order, compute running CVE totals
    running_total = cve_total_baseline
    for audit_id in audit_list:
        status = ''

        # Get the audit details
        sql = f"""SELECT content,create_time FROM wr_studio_harboraudit WHERE id = ?"""
        hb_audit = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()
        if not hb_audit:
            print(f"ERROR:NO AUDIT:{audit_id}")
            continue
        content = hb_audit['content']
        date = hb_audit['create_time'][:10]

        if audit_id in skip_id_dict:
            status = "SKIP_AUDIT"
        elif audit_id in date_skip_list:
            status = "SKIP_OLDER_DUPLICATE_DATE"

        # Fetch Total number of CVEs of the audit
        sql = f"""SELECT count(*) FROM wr_studio_harborvulnerability WHERE harboraudit_id = ?"""
        cve_total = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()[0]

        # Fetch the total number of artifacts in this audit
        sql = f"""SELECT count(*) FROM wr_studio_harborartifact WHERE harboraudit_id = ? AND (selected_top = True OR audit_status = 1)"""
        artifact_total = SQL_EXECUTE(cur, sql, (audit_id,)).fetchone()[0]

        # Fetch CVEs of Product total
        sql = f"""SELECT count(*) FROM wr_studio_cveoftheday WHERE harboraudit_id = ? AND content = ? AND is_added = True"""
        cor_total = SQL_EXECUTE(cur, sql, (audit_id,audit_content,)).fetchone()[0]

        # Fetch resolved CVEs of Product total
        sql = f"""SELECT count(*) FROM wr_studio_cveoftheday WHERE harboraudit_id = ? AND content = ? AND is_added = False"""
        ror_total = SQL_EXECUTE(cur, sql, (audit_id,audit_content,)).fetchone()[0]

        # Compute the running total (if not skipped)
        new_running_total = running_total - ror_total + cor_total
        if not status:
            running_total = new_running_total
            if running_total != cve_total:
                status = f"(Offeset = {cve_total-running_total})"
        else:
            new_running_total = f"({running_total})"

        print(f"#{audit_id:4} {cve_total:7} {new_running_total:>8} {cor_total:7} {ror_total:7} {date} {artifact_total:9} {content} {status}")


#################################
# scan_cve_sla
# ./bin/wr_studio/srtool_cves.py --call-scan-cve-sla 278 735

ENABLE_REPO_MAPPINGS = False

# This original version uses db recs with indexes, indtead of dictionaries
def scan_cve_sla_idx(aud_id1,passed_conn=None):
    print('passed audit id to function scan_cve_sla is-->',aud_id1)
    if not passed_conn:
        conn = SQL_CONNECT()    # (column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    severity_check = {'critical':10, 'high':9, 'medium':8 , 'low':7, 'unknown':6}
    cve_summary ={}
    sql1 =  """SELECT create_time from wr_studio_harboraudit where id = ?;"""
    audit_date1 = (SQL_EXECUTE(cur, sql1,(aud_id1,)).fetchall())[0][0]
    audit_date2 = audit_date1.split()[0]    # Isolate date from time
    audit_date = (datetime.strptime(audit_date2, '%Y-%m-%d')).date()
    print(f'1.date of audit {aud_id1} is--->',audit_date)

    sql2 = """SELECT id FROM wr_studio_harborartifact WHERE harboraudit_id = ?;"""
    artifact_list_sql = SQL_EXECUTE(cur, sql2,(aud_id1,)).fetchall()
    artifact_list = []
    for i1 in artifact_list_sql:
        artifact_list.append(i1[0])
    print('2.first 5 artifact ids--->',artifact_list[0:5])
    print('3.len of artifact_list-->',len(artifact_list))

    for artifact in artifact_list:
        sql3 = """SELECT repo.name,art.tag FROM
        wr_studio_harborrepository repo inner join wr_studio_harborartifact art on repo.id = art.harborrepository_id WHERE art.id = ?;"""
        repo_name,tag = (SQL_EXECUTE(cur, sql3,(artifact,)).fetchall())[0]
        #print(repo_name,tag)

        sql4 = """SELECT name,severity FROM wr_studio_harborvulnerability WHERE harborartifact_id = ?;"""
        vuln_list_sql = SQL_EXECUTE(cur, sql4,(artifact,)).fetchall()
        vuln_list = []
        for i2 in vuln_list_sql:
            vuln_list.append(i2)
        #print('4.len of vuln list-->',len(vuln_list))

# CVE-2009-5080]: R=wr-studio-product/wrlinux-builder
        for vuln in vuln_list:
            vuln_name,severity = vuln
#
            if (vuln_name == 'CVE-2009-5080') and (repo_name == 'wr-studio-product/wrlinux-builder'):
                print(f"4:FOUND:[{aud_id1}]CVE-2009-5080:wr-studio-product/wrlinux-builder:{tag}")

            age_key = f"{vuln_name}|{repo_name}|{tag}"
            if age_key in cve_summary:
                exist_severity = cve_summary[age_key]
                if severity_check[exist_severity.lower()] >= severity_check[severity.lower()]:
                    pass
                else:
                    cve_summary[age_key] = severity
            else:
                cve_summary[age_key] = severity

    print('5.len of cve summary dictionary-->',len(cve_summary))
    sql5 = """SELECT name,repo_name,tag,id FROM wr_studio_cvesla;"""
    exist_records_sql = SQL_EXECUTE(cur, sql5,()).fetchall()
    exist_records = {}
    for rec in exist_records_sql:
        name1,repo_name1,tag1,id1 = rec
        exist_records[f"{name1}|{repo_name1}|{tag1}"] = id1

    # Fetch SLA day counts
    add_days = {
        'critical' : int(srtsetting_get(conn,'HARBOR_SLA_CRITICAL','15',False)),
        'high'     : int(srtsetting_get(conn,'HARBOR_SLA_HIGH','35',False)),
        'medium'   : int(srtsetting_get(conn,'HARBOR_SLA_MEDIUM','60',False)),
        'low'      : int(srtsetting_get(conn,'HARBOR_SLA_LOW','100',False)),
        'unknown'  : int(srtsetting_get(conn,'HARBOR_SLA_UNKNOWN','100',False)),
    }

    insert_recs_list = []
    update_recs_list = []

    mode = 0
    for key1 in cve_summary:
        cve_name,repo_name,tag = key1.split('|')

#
        if (cve_name == 'CVE-2009-5080') and (repo_name == 'wr-studio-product/wrlinux-builder'):
            print(f"5:FOUND:[{aud_id1}]CVE-2009-5080:wr-studio-product/wrlinux-builder:{tag}")

        severity = cve_summary[key1]
        tag_base = ''

        if key1 in exist_records:
            #for updating existing record
            id1 = exist_records[key1]
            params = (audit_date,id1)
            update_recs_list.append(params)
        else:
            #for creating new record
            cve_id1 = ''
            nvd_date1 = ''
            if cve_name.startswith('CVE'):
                sql6 = """SELECT id,publisheddate FROM orm_cve WHERE name = ?;"""
                cve_rec_sql = SQL_EXECUTE(cur, sql6,(cve_name,)).fetchall()
                if len(cve_rec_sql) > 0:
                    cve_id1,nvd_date1 = cve_rec_sql[0]
                else:
                    cve_id1 = None
                    nvd_date1 = None
            else:
                cve_id1 = None
                nvd_date1 = audit_date
            dys = add_days.get(severity.lower(),100)
            sla_date1 = audit_date + timedelta(days = dys)
            params = (cve_name,cve_id1,repo_name,tag,tag_base,severity,nvd_date1,audit_date,sla_date1,audit_date,mode)
            insert_recs_list.append(params)

    print('6.len of insert_recs_list-->',len(insert_recs_list))
    print('7.len of update_recs_list-->',len(update_recs_list))

    if len(insert_recs_list) > 0:
        SQL_BATCH_WRITE(cur,
                "wr_studio_cvesla",
                insert_recs_list,
                fields=["name","cve_id","repo_name","tag","tag_base","severity","nvd_date","repo_date","sla_date","update_date","mode"],
                override_values=None,
                )
        SQL_COMMIT(conn)
        print('8.data inserted in cvesla table')

    if len(update_recs_list) > 0:
        SQL_BATCH_UPDATE(cur,
                "wr_studio_cvesla",
                update_recs_list,
                set_field=["update_date"],
                where_field=["id"],
                )
        SQL_COMMIT(conn)
        print('9.data updated in cvesla table')

    # Close the connection
    if not passed_conn:
        SQL_CLOSE_CONN(conn)

def scan_cve_sla(aud_id1,passed_conn=None):
    print('passed audit id to function scan_cve_sla is-->',aud_id1)
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    severity_check = {'critical':10, 'high':9, 'medium':8 , 'low':7, 'unknown':6}
    cve_summary ={}
    sql1 =  """SELECT create_time from wr_studio_harboraudit where id = ?;"""
    audit_date1 = (SQL_EXECUTE(cur, sql1,(aud_id1,)).fetchall())[0]['create_time']
    audit_date2 = audit_date1.split()[0]    # (2023-02-26 17:20:28.991761)
    audit_date = (datetime.strptime(audit_date2, '%Y-%m-%d')).date()
    print(f'1.date of audit {aud_id1} is--->',audit_date)

    sql2 = """SELECT id FROM wr_studio_harborartifact WHERE harboraudit_id = ?;"""
    artifact_list_sql = SQL_EXECUTE(cur, sql2,(aud_id1,)).fetchall()
    artifact_list = []
    for i1 in artifact_list_sql:
        artifact_list.append(i1['id'])
    print('2.first 5 artifact ids--->',artifact_list[0:5])
    print('3.len of artifact_list-->',len(artifact_list))

    for artifact in artifact_list:
        sql3 = """SELECT repo.name,art.tag FROM
        wr_studio_harborrepository repo inner join wr_studio_harborartifact art on repo.id = art.harborrepository_id WHERE art.id = ?;"""
        hb_obj = (SQL_EXECUTE(cur, sql3,(artifact,)).fetchall())[0]
        repo_name = hb_obj['name'].replace('3rdparty/','external/')
        tag = hb_obj['tag']
        #print(repo_name,tag)

#        print(f"FOO:{hb_obj['name']}:{repo_name}")

        sql4 = """SELECT name,severity FROM wr_studio_harborvulnerability WHERE harborartifact_id = ?;"""
        vuln_list_sql = SQL_EXECUTE(cur, sql4,(artifact,)).fetchall()
        vuln_list = []
        for i2 in vuln_list_sql:
            vuln_list.append([i2['name'],i2['severity']])
        #print('4.len of vuln list-->',len(vuln_list))

        for vuln in vuln_list:
            vuln_name,severity = vuln
            age_key = f"{vuln_name}|{repo_name}|{tag}"
            if age_key in cve_summary:
                exist_severity = cve_summary[age_key]
                if severity_check[exist_severity.lower()] >= severity_check[severity.lower()]:
                    pass
                else:
                    cve_summary[age_key] = severity
            else:
                cve_summary[age_key] = severity

    print('5.len of cve summary dictionary-->',len(cve_summary))
    sql5 = """SELECT name,repo_name,tag,id FROM wr_studio_cvesla;"""
    exist_records_sql = SQL_EXECUTE(cur, sql5,()).fetchall()
    exist_records = {}
    for rec in exist_records_sql:
#        name1,repo_name1,tag1,id1 = rec
        name1 = rec['name']
        repo_name1 = rec['repo_name']
        tag1 = rec['tag']
        id1 = rec['id']
        exist_records[f"{name1}|{repo_name1}|{tag1}"] = id1

    # Fetch SLA day counts
    add_days = {
        'critical' : int(srtsetting_get(conn,'HARBOR_SLA_CRITICAL','15')),
        'high'     : int(srtsetting_get(conn,'HARBOR_SLA_HIGH','35')),
        'medium'   : int(srtsetting_get(conn,'HARBOR_SLA_MEDIUM','60')),
        'low'      : int(srtsetting_get(conn,'HARBOR_SLA_LOW','100')),
        'unknown'  : int(srtsetting_get(conn,'HARBOR_SLA_UNKNOWN','100')),
    }

    insert_recs_list = []
    update_recs_list = []

    mode = 0
    for key1 in cve_summary:
        cve_name,repo_name,tag = key1.split('|')
        severity = cve_summary[key1]
        tag_base = ''

        # NOTE: hack to map Ingest repo names to Develop
        key1_develop = f"{cve_name}|{repo_name.replace('3rdparty/','external/')}|{tag}"

        if key1 in exist_records:
            #for updating existing record
            id1 = exist_records[key1]
            params = (audit_date,id1)
            update_recs_list.append(params)
        elif key1_develop in exist_records:
            #for updating existing record
            id1 = exist_records[key1_develop]
            params = (audit_date,id1)
            update_recs_list.append(params)
        else:
            #for creating new record
            cve_id1 = ''
            nvd_date1 = ''
            if cve_name.startswith('CVE'):
                sql6 = """SELECT id,publisheddate FROM orm_cve WHERE name = ?;"""
                cve_rec_sql = SQL_EXECUTE(cur, sql6,(cve_name,)).fetchone()
                if cve_rec_sql:
#                    cve_id1,nvd_date1 = cve_rec_sql[0]
                    cve_id1 = cve_rec_sql['id']
                    nvd_date1 = cve_rec_sql['publisheddate']
                else:
                    cve_id1 = None
                    nvd_date1 = None
            else:
                cve_id1 = None
                nvd_date1 = audit_date
            dys = add_days.get(severity.lower(),100)
            sla_date1 = audit_date + timedelta(days = dys)
            params = (cve_name,cve_id1,repo_name,tag,tag_base,severity,nvd_date1,audit_date,sla_date1,audit_date,mode)
            insert_recs_list.append(params)

    print('6.len of insert_recs_list-->',len(insert_recs_list))
    print('7.len of update_recs_list-->',len(update_recs_list))

    if len(insert_recs_list) > 0:
        SQL_BATCH_WRITE(cur,
                "wr_studio_cvesla",
                insert_recs_list,
                fields=["name","cve_id","repo_name","tag","tag_base","severity","nvd_date","repo_date","sla_date","update_date","mode"],
                override_values=None,
                )
        SQL_COMMIT(conn)
        print('8.data inserted in cvesla table')

    if len(update_recs_list) > 0:
        SQL_BATCH_UPDATE(cur,
                "wr_studio_cvesla",
                update_recs_list,
                set_field=["update_date"],
                where_field=["id"],
                )
        SQL_COMMIT(conn)
        print('9.data updated in cvesla table')

    # Close the connection
    if not passed_conn:
        SQL_CLOSE_CONN(conn)


# id1 is start audit id(older audit), id2 is end audit id(newer audit)
# ./bin/wr_studio/srtool_cves.py --call-scan-cve-sla 278 725
def call_scan_cve_sla(id1,id2,passed_conn=None):
    print('call_scan_cve_sla function called with-->',id1,id2)
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    CVEAGE_LOG = 'cveage_trace.log'

    current_first_sla_audit = int(srtsetting_get(conn,'HARBOR_SLA_FIRST','99999'))
    current_last_sla_audit = int(srtsetting_get(conn,'HARBOR_SLA_LAST','0'))

    # Fetch control file for repo|basetag remappings table
    repo_key_remap = {}
    audit_keep = {}
    audit_skip = {}
    file_fd = open('data/cve_age_repo_map.csv', 'r')
    for line in file_fd.readlines():
        line = line.strip()
        if (not line) or ('#' == line[0]):
            continue
        action,a,b = line.split(',')
        if ('MAP' == action) and ENABLE_REPO_MAPPINGS:
            repo_key_remap[a] = b
        if 'KEEP' == action:
            audit_keep[int(a)] = True
        if 'SKIP' == action:
            audit_skip[int(a)] = True
    if verbose:
        print(f"KEEP={audit_keep.keys()}")
        print(f"SKIP={audit_skip.keys()}")

    # Log the registered audits
    fd = open(CVEAGE_LOG, 'a')

    sql1 = """SELECT id,name FROM wr_studio_harboraudit WHERE (id BETWEEN ? AND ?) AND content = ? ORDER BY id;"""
    audit_ids_sql = SQL_EXECUTE(cur, sql1,(id1,id2,'conductor')).fetchall()
    audit_ids = []
    for i,audit in enumerate(audit_ids_sql):
        audit_id = int(audit[ORM.WR_STUDIO_HARBORAUDIT_ID])
        audit_name = audit[ORM.WR_STUDIO_HARBORAUDIT_NAME]
        if audit_id in audit_keep:
            action = "ADD_KEEP"
        elif audit_id in audit_skip:
            action = "SKIP_AUDIT"
        elif not "results_develop" in audit_name:
            action = "SKIP_TRIVY"
        else:
            action = "ADD_AUDIT"
        msg = f"{i:3}) {action} [{audit_id}]: {audit_name}"
        fd.write(msg + '\n')
        if verbose: print(msg)
        if action.startswith('ADD'):
            audit_id = int(audit[ORM.WR_STUDIO_HARBORAUDIT_ID])
            audit_ids.append(audit_id)
    print('audit ids in selected range-->',audit_ids)

    progress_set_max( (len(audit_ids)*2) + 1)
    for audit_id in audit_ids:
        scan_cve_sla(audit_id,passed_conn=conn)
        progress_show(f"Audit #{audit_id} scan")
        if audit_id < current_first_sla_audit:
            srtsetting_set(conn,'HARBOR_SLA_FIRST',str(audit_id))
        if audit_id > current_last_sla_audit:
            srtsetting_set(conn,'HARBOR_SLA_LAST',str(audit_id))

    fd.close()
    if not passed_conn:
        SQL_CLOSE_CONN(conn)

#################################
# cve_sla_develop
# ./bin/wr_studio/srtool_cves.py --cve-sla-mode 1
#
# Convert raw CVE Age data (mode 0) into audit display
# compatible data
#

def cve_sla_mode(mode=1,passed_conn=None):
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    # [CVE-2009-5155]: R=wr-studio-product/external/docker.io/mongo, T=4.4.6, B=, M=0
    check_cve = 'CVE-2009-5155'
    check_repo = 'wr-studio-product/external/docker.io/mongo'
    check_tag = '4.4.6'

    if verbose: print(f"CVE_SLA_MODE:{mode}")

    # Fetch managed tag repos
    managed_tags = {}
    for managed_tag in srtsetting_get(conn,'STUDIO_MANAGED_TAGS_SETS','').split('\n'):
        managed_tags[managed_tag] = True
    def get_tag_base(repo_name,tag):
        is_managed = False
        for managed_tag in managed_tags:
            if (not managed_tag) or ('#' == managed_tag[0]):
                continue
            repo_name_str,tag_base_str = managed_tag.split(':')
            if (repo_name == repo_name_str):
                is_managed = True
                if tag.startswith(tag_base_str):
                    return(tag_base_str)
        # If managed but no tag_base, error
        if is_managed:
            print(f"WARNING:Managed '{repo_name}' but tag-base not found '{tag}'")
            tag_base = tag[:tag.find('-')+1]
            managed_tags[f"{repo_name}:{tag_base}"] = True
            return(tag_base)
        # Non-managed repos have an empty tag_base to join all tags
        return('')

    # Fetch manifest for mode
    if verbose: print(f"CVE_SLA_MODE:Manifest")
    audit_manifest = {}
    if 1 == mode:
        audit_content = 'develop'
    elif 2 == mode:
        audit_content = 'develop'
    else:
        print(f"ERROR: unknown mode '{mode}'")
        exit(1)
    cmnd = ['./bin/wr_studio/srtool_harbor.py','--list-conductor-manifest',audit_content]
    exec_returncode,exec_stdout,exec_stderr = execute_process(cmnd)
    for line in exec_stdout.split('\n'):
        # "REPO:wr-studio-product/ota/backend=ota-reverse-proxy-1.1"
        if not line.startswith('REPO:'):
            continue
        line = line[5:]
        repo_name,tag = line.split('=')
        tag_base = get_tag_base(repo_name,tag)
        key = f"{repo_name}|{tag_base}"
        audit_manifest[key] = True
    print(f"audit_manifest_count={len(audit_manifest)}")

    # Fetch control file for repo|basetag remappings table
    if verbose: print(f"CVE_SLA_MODE:Control file")
    repo_key_remap = {}
    file_fd = open('data/cve_age_repo_map.csv', 'r')
    for line in file_fd.readlines():
        line = line.strip()
        if (not line) or ('#' == line[0]):
            continue
        action,before,after = line.split(',')
        if ENABLE_REPO_MAPPINGS and (action == 'MAP'):
            repo_key_remap[before] = after
    print(f"repo_remap_count={len(repo_key_remap)}")

    # Order by name then repo_date (oldest for repo multiple tags wins)
    sql = f"""SELECT * FROM wr_studio_cvesla WHERE mode = 0 ORDER BY name,repo_name,repo_date;"""

    insert_recs_list = []               # Mode 1 records to add
    update_recs_list = []               # Mode 1 records to update

    unused_recs_dict = {}               # Unused Manifest repos
    for key in audit_manifest:
        unused_recs_dict[key] = True
    found_recs_dict = {}                #
    duplicate_cves_dict = {}            #
    skip_keys = {}
    record_count = 0                    # Mode 0 records read
    for i,sla in enumerate(SQL_EXECUTE(cur, sql,).fetchall()):
        record_count += 1
        if 0 == (i % 100):
            print(f"{i:6}...",end='\r')

#
        # Check test
        if (check_cve == sla['name']) and (check_repo == sla['repo_name']) and (check_tag == sla['tag']):
            print(f"CHECK1a:{check_cve}:{check_repo}:{check_tag}")
        if 5639077 == int(sla['id']):
            print(f"CHECK1b:{check_cve}:{check_repo}:{check_tag}")
        # CHECK1b:CVE-2009-5155:wr-studio-product/external/docker.io/mongo:4.4.6

        # Test the repo|tag_base first
        tag_base = get_tag_base(sla['repo_name'],sla['tag'])
        repo_tag_key = f"{sla['repo_name']}|{tag_base}"
        if repo_tag_key in repo_key_remap:
            repo_tag_key = repo_key_remap[repo_tag_key]
#        if repo_tag_key in duplicate_recs_dict:
#            continue
        # INVALID: older audit may have used a different manifest than the current
        if False:
            if not repo_tag_key in audit_manifest:
                skip_keys[repo_tag_key] = True
                continue
        if repo_tag_key in unused_recs_dict:
            del unused_recs_dict[repo_tag_key]
        found_recs_dict[repo_tag_key] = True

        # Test the CVE|repo_tag second
        if 1 == mode:
            cve_key = f"{sla['name']}|{repo_tag_key}"
        elif 2 == mode:
            cve_key = f"{sla['name']}"
        if cve_key in duplicate_cves_dict:
            continue
        duplicate_cves_dict[cve_key] = True
        # Capture the CVE
        params = (
            sla['name'],
            sla['cve_id'],
            sla['repo_name'] if (1==mode) else '',
            sla['tag'] if (1==mode) else '',
            tag_base if (1==mode) else '',
            sla['severity'],
            sla['nvd_date'],
            sla['repo_date'],
            sla['sla_date'],
            sla['update_date'],
            mode)

        # Insert or update
        sql = """SELECT id FROM wr_studio_cvesla WHERE name = ? AND repo_name = ? AND tag = ? AND mode = ?;"""
        exist_record_cve_age = SQL_EXECUTE(cur, sql,(sla['name'],sla['repo_name'],sla['tag'],1)).fetchone()
        if not exist_record_cve_age:
            insert_recs_list.append(params)
        else:
            update_recs_list.append(params)

        # Check test
        if (check_cve == sla['name']) and (check_repo == sla['repo_name']) and (check_tag == sla['tag']):
            print(f"CHECK9:{check_cve}:{check_repo}:{check_tag}")

    # Pre-delete existing mode entries
    if False:
        sql = 'DELETE FROM wr_studio_cvesla WHERE mode=?'
        SQL_EXECUTE(cur, sql, (mode,))
        SQL_COMMIT(conn)

    # Bulk insert new list
    print(f"Insert count = {len(insert_recs_list)}")
    if len(insert_recs_list) > 0:
        SQL_BATCH_WRITE(cur,
                "wr_studio_cvesla",
                insert_recs_list,
                fields=["name","cve_id","repo_name","tag","tag_base","severity","nvd_date","repo_date","sla_date","update_date","mode"],
                override_values=None,
                )
        SQL_COMMIT(conn)
    # Bulk update new list
    print(f"Update count = {len(update_recs_list)}")
    if len(update_recs_list) > 0:
        # No updates at this time
        pass

    HARBOR_SLA_FIRST = int(srtsetting_get(conn,'HARBOR_SLA_FIRST','0'))
    HARBOR_SLA_LAST = int(srtsetting_get(conn,'HARBOR_SLA_LAST','0'))
    if verbose:
        print("AUDIT_MANIFEST:")
        for i,key in enumerate(sorted(audit_manifest)):
            print(f"M [{i:3}] {key}")
        print("INCLUDE_KEYS:")
        for i,key in enumerate(sorted(found_recs_dict)):
            print(f"I [{i:3}] {key}")
        print("SKIP_KEYS:")
        for i,key in enumerate(sorted(skip_keys)):
            print(f"O [{i:3}] {key}")
        print("UNUSED_KEYS:")
        for i,key in enumerate(sorted(unused_recs_dict)):
            msg = ''
            repo_name = re.sub('\|.*','',key)
            sql = """SELECT * FROM wr_studio_harborrepository WHERE name = ? AND harboraudit_id = ?;"""
            hb_repo = SQL_EXECUTE(cur, sql, (repo_name,HARBOR_SLA_LAST)).fetchone()
            if hb_repo:
                sql = """SELECT * FROM wr_studio_harborartifact WHERE harborrepository_id = ?;"""
                hb_artifacts = SQL_EXECUTE(cur, sql, (hb_repo['id'],))
                if hb_artifacts:
                    vul_count = 0
                    for hb_artifact in hb_artifacts:
                        sql = """SELECT id FROM wr_studio_harborvulnerability WHERE harborartifact_id = ?;"""
                        vulnerabilities = SQL_EXECUTE(cur, sql, (hb_artifact['id'],))
                        if vulnerabilities:
                            vul_count += len(vulnerabilities.fetchall())
                    msg = f"V={vul_count:5}"
                else:
                    #print(f"No Artifacts '{repo_name}',{AUDIT_ID}")
                    msg = 'NO_ARTS'
            else:
                #print(f"No Repo '{repo_name}',{AUDIT_ID}")
                msg = 'NO_REPO'
            print(f"U [{i:3}] {msg}:{key}")

    print("")
    print(f"HARBOR_SLA_FIRST   = {HARBOR_SLA_FIRST}")
    print(f"HARBOR_SLA_LAST    = {HARBOR_SLA_LAST}")
    print(f"Audit_manifest     = {len(audit_manifest)}")
    print(f"In  records        = {record_count}")
    print(f"Out records        = {len(insert_recs_list)}")
    print(f"Included repos     = {len(found_recs_dict)}")
    print(f"Mapped   old repos = {len(repo_key_remap)}")
    print(f"Unmapped old repos = {len(skip_keys)}")
    print(f"Missed audit repos = {len(unused_recs_dict)}")

    if not passed_conn:
        SQL_CLOSE_CONN(conn)

#################################
# cve_sla_update_all
#

def cve_sla_update_all(range):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    if 'default' == range:
        current_last_sla_audit = int(srtsetting_get(conn,'HARBOR_SLA_LAST','0'))
        # Find latest audit ID
        sql1 = """SELECT id,name FROM wr_studio_harboraudit WHERE content = ? ORDER BY id DESC;"""
        hb_audit_top = SQL_EXECUTE(cur, sql1,('conductor',)).fetchone()
        latest_audit_id = hb_audit_top['id']
        print(f"CVE_SLA_UPDATE_ALL:DEFAULT:CurrentLast={current_last_sla_audit}:Latest={latest_audit_id}")
    else:
        ranges = range.split(',')
        current_last_sla_audit = int(ranges[0])
        latest_audit_id = int(ranges[1])
        print(f"CVE_SLA_UPDATE_ALL:CurrentLast={current_last_sla_audit}:Latest={latest_audit_id}")

    # Scan CVE ages for audit range
    print(f"=== 1. Scan CVE ages ({current_last_sla_audit} to {latest_audit_id}) ===")
    call_scan_cve_sla(current_last_sla_audit,latest_audit_id,passed_conn=conn)

    # Parse CVE ages for Vulnerability tables
    print("=== 2. Convert CVE ages to CVEs page format ===")
    progress_show("Merge CVE Ages")
    cve_sla_mode(1,passed_conn=conn)

    # Update audits with CVE age info
    print("=== 3. Apply CVE ages to recent audits ===")
    if current_last_sla_audit < latest_audit_id:
        sql1 = """SELECT id,name FROM wr_studio_harboraudit WHERE (id BETWEEN ? AND ?) AND content = ? ORDER BY id;"""
        audit_ids_sql = SQL_EXECUTE(cur, sql1,(current_last_sla_audit,latest_audit_id,'conductor')).fetchall()
        for i,hb_audit in enumerate(audit_ids_sql):
            progress_show(f"Audit #{hb_audit['id']} ages update")
            print(f" == [{i:4}] {hb_audit['id']}: {hb_audit['name']} ==")
            cve_sla_apply_audit(hb_audit['id'],is_dry_run=False,passed_conn=conn)
    else:
        _log(f"WARNING: CVE_SLA_UPDATE_ALL:SKIP AGE UPDATE:{current_last_sla_audit} to {latest_audit_id}")
    SQL_CLOSE_CONN(conn)

def cve_sla_delete():
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # Delete all Vulnerability SLA links
    update_vuls_list = []
    sql = """select id,cve_sla_id from wr_studio_harborvulnerability WHERE cve_sla_id IS NOT NULL """
    for i,vul in enumerate(SQL_EXECUTE(cur, sql,).fetchall()):
        params = (None,vul['id'])
        update_vuls_list.append(params)
    print(f"* Delete {len(update_vuls_list)} Vulnerability links to CveSla")
    SQL_BATCH_UPDATE(cur,
            "wr_studio_harborvulnerability",
            update_vuls_list,
            set_field=["cve_sla_id"],
            where_field=["id"],
            )

    # Delete existing mode entries
    print(f"* Erase CveSla table content")
    sql = 'DELETE FROM wr_studio_cvesla;'
    SQL_EXECUTE(cur, sql, )
    SQL_COMMIT(conn)

    # Clear the SLA covered audit log
    try:
        os.remove(CVEAGE_LOG)
    except:
        pass

    # Reset the SLA covered range
    srtsetting_set(conn,'HARBOR_SLA_FIRST','99999')
    srtsetting_set(conn,'HARBOR_SLA_LAST','0')

    SQL_COMMIT(conn)
    SQL_CLOSE_CONN(conn)

def cve_sla_dump():
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    log_name = 'cve_age_table.txt'
    log_file = open(log_name, 'w')
    # Dump all CVE SLA records to text file
    update_vuls_list = []
    sql = """select name,repo_name,tag from wr_studio_cvesla ORDER BY name,repo_name,tag """
    for i,sla in enumerate(SQL_EXECUTE(cur, sql,).fetchall()):
        if 0 == (i % 1024):
            print(f"{i:5}\t",end='')
        log_file.write(f"{i},{sla['name']},{sla['repo_name']},{sla['tag']}\n" )
    print('')
    log_file.close()
    SQL_CLOSE_CONN(conn)
    print(f"The CveSla table {i} was dumped to '{log_name}'")

#################################
# merge_ingest
#
# Theory of Ingest/Harbor mapping:
# 1) Remove product and "external","3rdparty" prefixes
# 2) Do hardcoded remapping via 'ingest_hard_mappings'
# 3) Attempt a full mapping of current paths
# 4) Attept a mapping on the last path names
# 5) Give up as no mapping
#
#
#

def merge_ingest(audit_id):
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    # Ingest forced mappings from Ingest to Harbor
    ingest_hard_mappings = {
        'google/cloud-sdk' : 'cloudsdktool',
        'jenkins/jenkins' : 'wrstudio-jenkins',
        'ghcr.io/zalando/spilo-15' : 'wrstudio/opensrc/util/spilo',
        'cgr.dev/chainguard/busybox' : 'ghcr.io/distroless/busybox',
        'registry.gitlab.com/gitlab-org/cloud-native/mirror/images/busybox' : 'registry.gitlab.com/gitlab-org/cloud-native/mirror/images/busybox',
        'busybox' : 'docker.io/busybox',
    }

    #
    # Find and load the latest INGEST audit
    #

    artifact_ingest_dict = {}
    sql = f"""SELECT * FROM wr_studio_harboraudit WHERE (name LIKE '%' || ? || '%') ORDER BY id DESC;"""
    hb_audit_ingest = SQL_EXECUTE(cur, sql,('results_ingest',)).fetchone()
    if not hb_audit_ingest:
        print("ERROR:missing ingest audit")
        exit(1)
    else:
        if verbose: print(f"== Ingest audit:{hb_audit_ingest['id']}")

    # Preload the INGEST repo name,id
    hb_ingest_repo_dict = {}
    sql = """SELECT id,name FROM wr_studio_harborrepository WHERE harboraudit_id = ?;"""
    for hb_ingest_repo in SQL_EXECUTE(cur, sql, (hb_audit_ingest['id'],)).fetchall():
        ingest_mapped_name = hb_ingest_repo['name'].replace('wr-studio-ingest/','').replace('3rdparty/','').replace('external/','')
        mapped_names = [ingest_mapped_name,os.path.basename(ingest_mapped_name)]
        if ingest_mapped_name in ingest_hard_mappings:
            # Add mapped path into dictionary for exceptions
            mapped_names.append(ingest_hard_mappings[ingest_mapped_name])
        hb_ingest_repo_dict[hb_ingest_repo['id']] = mapped_names
    if verbose:
        for ingest_match in hb_ingest_repo_dict:
            print(f"FOO:INGEST:{hb_ingest_repo_dict[ingest_match]}")

    # Now load the AUDIT artifacts,id
    sql = """SELECT id,harborrepository_id,tag FROM wr_studio_harborartifact WHERE harboraudit_id = ?;"""
    for i,artifact_ingest in enumerate(SQL_EXECUTE(cur, sql,(hb_audit_ingest['id'],)).fetchall()):
        artifact_ingest_names = hb_ingest_repo_dict[artifact_ingest['harborrepository_id']]
        for name in artifact_ingest_names:
            artifact_ingest_dict[name] = [artifact_ingest['id'],artifact_ingest['tag']]
#        print(f"I[{i:3}]:{artifact_ingest_names}")
    if verbose: print(f"== Ingest audit artifact count:{len(artifact_ingest_dict)}")

    #
    # Iterate the audit
    #

    # Preload the AUDIT repo name,id
    hb_audit_repo_dict = {}
    sql = """SELECT id,name FROM wr_studio_harborrepository WHERE harboraudit_id = ?;"""
    for i,hb_audit_repo in enumerate(SQL_EXECUTE(cur, sql, (audit_id,)).fetchall()):
        hb_audit_repo_dict[hb_audit_repo['id']] = hb_audit_repo['name'].replace('wr-studio-product/','').replace('external/','')
#        print(f"R[{i:3}]:{hb_audit_repo['id']}:{hb_audit_repo['name']}")

    def calculate_cve_counts(audit_test_id):
        # Gather the repo CVE statistics
        critical_count = 0
        high_count = 0
        medium_count = 0
        low_count = 0
        unknown_count = 0
        sql = """SELECT severity FROM wr_studio_harborvulnerability WHERE harborartifact_id = ?;"""
        for hb_vul in SQL_EXECUTE(cur, sql,(audit_test_id,)).fetchall():
            severity = hb_vul['severity']
            if 'Critical' == severity:
                critical_count += 1
            elif 'High' == severity:
                high_count += 1
            elif 'Medium' == severity:
                medium_count += 1
            elif 'Low' == severity:
                low_count += 1
            else:
                unknown_count += 1
        return critical_count,high_count,medium_count,low_count,unknown_count

    # Load the AUDIT artifacts, and compare to INGEST
    sql = """SELECT id,harborrepository_id FROM wr_studio_harborartifact WHERE harboraudit_id = ?;"""
    artifact_list_sql = SQL_EXECUTE(cur, sql,(audit_id,)).fetchall()
    if verbose: print(f"== Audit artifact count:{len(artifact_list_sql)}")
    artifact_list = []
    for i,hb_art in enumerate(artifact_list_sql):
#        print(f"A[{i:3}]:{hb_art['harborrepository_id']}")
        hb_repo_name = hb_audit_repo_dict[hb_art['harborrepository_id']]
        hb_repo_name_short = os.path.basename(hb_repo_name)
        # Try matching full path
        if hb_repo_name in artifact_ingest_dict:
            ingest_artifact_id,ingest_artifact_tag = artifact_ingest_dict[hb_repo_name]
        elif hb_repo_name_short in artifact_ingest_dict:
            # Try matching last path
            ingest_artifact_id,ingest_artifact_tag = artifact_ingest_dict[hb_repo_name_short]
        else:
            # No match
            if verbose: print(f"INGEST:NOMATCH:{hb_repo_name}:{hb_repo_name_short}:")
            continue
        # Gather the repo CVE statistics
        critical_count_audit,high_count_audit,medium_count_audit,low_count_audit,unknown_count_audit = calculate_cve_counts(hb_art['id'])
        vul_count_audit = critical_count_audit + high_count_audit + medium_count_audit + low_count_audit + unknown_count_audit
        critical_count_ingest,high_count_ingest,medium_count_ingest,low_count_ingest,unknown_count_ingest = calculate_cve_counts(ingest_artifact_id)
        vul_count_ingest = critical_count_ingest + high_count_ingest + medium_count_ingest + low_count_ingest + unknown_count_ingest
        # Add this new sum
        if (vul_count_ingest < vul_count_audit) or (critical_count_ingest < critical_count_audit) or (high_count_ingest < high_count_audit):
            ingest_status = '<'
        elif (vul_count_ingest > vul_count_audit) or (critical_count_ingest > critical_count_audit):
            ingest_status = '>'
        else:
            ingest_status = '='
        ingest = f"{ingest_artifact_id}|{ingest_status}|{ingest_artifact_tag}|{critical_count_ingest}|{high_count_ingest}|{medium_count_ingest}|{low_count_ingest}|{unknown_count_ingest}"
        params = (ingest,hb_art['id'])
        artifact_list.append(params)
#        if verbose: print(f"* [{audit_ingest_id}] {hb_repo_name}:{ingest}")

    # Bulk insert new list
    if verbose: print(f"== Audit artifact match:{len(artifact_list)}")
    if len(artifact_list) > 0:
        SQL_BATCH_UPDATE(cur,
                "wr_studio_harborartifact",
                artifact_list,
                set_field=["ingest",],
                where_field=["id"],
                )
        SQL_COMMIT(conn)

    if False:
        for ingest,hb_art_id in artifact_list:
            sql = """SELECT ingest FROM wr_studio_harborartifact WHERE id = ?;"""
            hb_art = SQL_EXECUTE(cur, sql,(hb_art_id,)).fetchone()
            print(f"UPDATED:{hb_art_id}:{ingest} => '{hb_art['ingest']}'")


    SQL_CLOSE_CONN(conn)

#################################
# fix_upper_severity
#

def fix_upper_severity():
    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    up_critical = 0
    audit_list = []
    for key in ('CRITICAL'):
        sql = """select severity,harboraudit_id from wr_studio_harborvulnerability"""
        for i,vul in enumerate(SQL_EXECUTE(cur, sql,).fetchall()):
            if 0 == (i % 1000):
                print(f"{i:6}...",end='\r')

            if "CRITICAL" == vul['severity'].strip():
                up_critical += 1
                audit_list.append(vul['harboraudit_id'])

    print(f"up_critical={up_critical}")
    print(f"audit_list={audit_list}")


#################################
# cve_sla_apply_audit
#

# Managed tag support
class Manage_Tags():
    # init method or constructor
    def __init__(self, conn, column_names=True):
        self.managed_tags = {}
        for managed_tag in srtsetting_get(conn,'STUDIO_MANAGED_TAGS_SETS','',column_names).split('\n'):
            self.managed_tags[managed_tag] = True
    def get_tag_base(self,repo_name,tag):
        is_managed = False
        for managed_tag in self.managed_tags:
            if (not managed_tag) or ('#' == managed_tag[0]):
                continue
            repo_name_str,tag_base_str = managed_tag.split(':')
            if (repo_name == repo_name_str):
                is_managed = True
                if tag.startswith(tag_base_str):
                    return(tag_base_str)
        return ''

# See if we can match all audit Vul to a CveSla record
# ./bin/wr_studio/srtool_cves.py --check-cveage-match 725
def cve_sla_apply_audit(audit_id,is_dry_run=False,passed_conn=None):
    if not passed_conn:
        conn = SQL_CONNECT(column_names=True)
    else:
        conn = passed_conn
    cur = SQL_CURSOR(conn)

    manage_tags = Manage_Tags(conn)

    age_table = {}
    age_repo_table = {}
    sql = f"""SELECT * FROM wr_studio_cvesla WHERE mode = 1 ;"""
    for i,cvesla in enumerate(SQL_EXECUTE(cur, sql,).fetchall()):
        key = f"{cvesla['name']}|{cvesla['repo_name']}|{cvesla['tag_base']}"
        age_table[key] = cvesla['id']
        key = f"{cvesla['repo_name']}|{cvesla['tag_base']}"
        age_repo_table[key] = True

    hb_artid2repoid_map = {}
    hb_repoid2name_map = {}

    sql = f"""SELECT * FROM wr_studio_harborvulnerability WHERE harboraudit_id = ? ORDER BY name;"""
    vul_count = 0
    found_count = 0
    found_repos = {}
    missing_repos = {}
    missing_ages = {}
    update_vuls_list = []
    for hb_vul in SQL_EXECUTE(cur, sql, (audit_id,)).fetchall():
        vul_count += 1
        vul_name = hb_vul['name']
        vul_id = hb_vul['id']
        vul_cve_sla_id = hb_vul['cve_sla_id']
        # Find repo_id
        hb_art_id = hb_vul['harborartifact_id']
        if not hb_art_id in hb_artid2repoid_map:
            sql = f"""SELECT * FROM wr_studio_harborartifact WHERE id = ?;"""
            hb_art = SQL_EXECUTE(cur, sql, (hb_art_id,)).fetchone()
            repo_id = hb_art['harborrepository_id']
            tag = hb_art['tag']
            hb_artid2repoid_map[hb_art_id] = [repo_id,tag]
        else:
            repo_id,tag = hb_artid2repoid_map[hb_art_id]
        # Find repo_name
        if not repo_id in hb_repoid2name_map:
            sql = f"""SELECT * FROM wr_studio_harborrepository WHERE id = ?;"""
            hb_repo = SQL_EXECUTE(cur, sql, (repo_id,)).fetchone()
            repo_name = hb_repo['name'].replace('3rdparty/','external/')
            hb_repoid2name_map[repo_id] = repo_name
        else:
            repo_name = hb_repoid2name_map[repo_id]

        tag_base = manage_tags.get_tag_base(repo_name,tag)

        repo_key = f"{repo_name}|{tag_base}"
        age_key = f"{vul_name}|{repo_key}"
        if age_key in age_table:
            found_count += 1
            found_repos[repo_key] = True

            # Update Vulnerability record if needed
            cve_sla_id = age_table[age_key]
            if cve_sla_id != vul_cve_sla_id:
                params = (cve_sla_id,vul_id)
                update_vuls_list.append(params)
        else:
            missing_repos[repo_key] = True
            missing_ages[age_key+f"||{tag}"] = True

    # What repos are totally absent?
    absent_repos = {}
    for repo_key in missing_repos:
        if not repo_key in found_repos:
            absent_repos[repo_key] = True
    recoverable_ages = 0
    for age_key in missing_ages:
        repo_key = age_key[age_key.find('|')+1:]
        if repo_key in absent_repos:
            recoverable_ages += 1

    # Bulk insert cve_sla updates
    if (not is_dry_run) and (len(update_vuls_list) > 0):
        SQL_BATCH_UPDATE(cur,
                "wr_studio_harborvulnerability",
                update_vuls_list,
                set_field=["cve_sla_id"],
                where_field=["id"],
                )
        SQL_COMMIT(conn)

    log_file = open('cve_age_match.log', 'w')
    print(f"update_vuls  = {len(update_vuls_list)}")
    print(f"vul_count    = {vul_count}")
    print(f"found_count  = {found_count}")
    print(f"missing ages = {len(missing_ages)} (recoverable={recoverable_ages})")
    for i,age_key in enumerate(missing_ages):
        log_file.write(f"AGE?:{age_key}\n" )
        if i < 10:
            print(f"  {age_key}")
    print(f"absent repos = {len(absent_repos)}")
    for i,repo_name in enumerate(absent_repos):
        log_file.write(f"MISS:{repo_name}\n" )
        if i < 10:
            print(f"  {repo_name}")
    print(f"found   repos= {len(found_repos)}")
    for i,repo_name in enumerate(found_repos):
        log_file.write(f"FIND:{repo_name}\n" )
        if i < 10:
            print(f"  {repo_name}")
    log_file.close()

    # Print error for progress
    if len(missing_ages):
        print(f"ERROR: missing ages = {len(missing_ages)} (recoverable={recoverable_ages})")

    if not passed_conn:
        SQL_CLOSE_CONN(conn)

# ./bin/wr_studio/srtool_cves.py --analyze-cveage-match 725
def analyze_cveage_match(audit_id_range):
    audit_id_range = audit_id_range.split(',')
    index_num = int(audit_id_range[0])
    audit_id_last = int(audit_id_range[1])
    audit_id_first = int(audit_id_range[2])
    if audit_id_first > audit_id_last:
        audit_id_first,audit_id_last = audit_id_last,audit_id_first

    conn = SQL_CONNECT(column_names=True)
    cur = SQL_CURSOR(conn)

    missing_ages = []
    log_file = open('cve_age_match.log', 'r')
    for line in log_file.readlines():
        line = line.strip()
        if (not line) or ('#' == line[0]):
            continue

        if line.startswith('AGE?:'):
            missing_ages.append(line[5:])

    for index in range(index_num,index_num+1):
        # AGE?:CVE-2009-5080|wr-studio-product/wrlinux-builder|
        cve_name,repo,tag_base = missing_ages[index].split('|')
        print(f"AGE:{cve_name},{repo},{tag_base}")

        print(f"AUDIT_IDS:{audit_id_first}:{audit_id_last}")
        is_found = False
        for audit_id in range(audit_id_first,audit_id_last+1):
            sql = """SELECT * FROM wr_studio_harboraudit WHERE id = ?;"""
            hb_audit = SQL_EXECUTE(cur, sql, (audit_id,))
            hb_audit = hb_audit.fetchone()
            if not hb_audit:
    #            print("SKIP:NO_AUDIT")
                continue
            if 'conductor' != hb_audit['content']:
    #            print(f"SKIP:NOT_CONDUCTOR ({hb_audit['content']})")
                continue

            if is_found:
                print(f"* V[{audit_id}]:{cve_name}")

            sql = """SELECT * FROM wr_studio_harborvulnerability WHERE harboraudit_id = ? AND name = ?;"""
            vulnerabilities = SQL_EXECUTE(cur, sql, (audit_id,cve_name,)).fetchall()
            if not vulnerabilities:
                pass
    #            print("  NOT FOUND")
                continue
            if not is_found:
                is_found = True
                print(f"* V[{audit_id}]:{cve_name}")
            for vulnerability in vulnerabilities:
                # Artifact
                sql = """SELECT * FROM wr_studio_harborartifact WHERE id = ?;"""
                artifact = SQL_EXECUTE(cur, sql, (vulnerability['harborartifact_id'],)).fetchone()
                # Repo
                sql = """SELECT * FROM wr_studio_harborrepository WHERE id = ?;"""
                repo = SQL_EXECUTE(cur, sql, (artifact['harborrepository_id'],)).fetchone()
                # Found
                print(f"  [{cve_name}]: R={repo['name']}, T={artifact['tag']}, P={vulnerability['package']}")

        print(f"\n* SLAs:{cve_name}")
        sql = """SELECT * FROM wr_studio_cvesla WHERE name = ?;"""
        slas = SQL_EXECUTE(cur, sql, (cve_name,)).fetchall()
        for sla in slas:
            print(f"  [{cve_name}]: R={sla['repo_name']}, T={sla['tag']}, B={sla['tag_base']}, M={sla['mode']}")

    SQL_CLOSE_CONN(conn)

#################################
# main loop
#

def main(argv):
    global srt_lx_remote
    global srt_lx_path
    global verbose
    global test
    global force_update
    global cmd_count
    global cmd_skip

    parser = argparse.ArgumentParser(description='srtool_cves.py: Generate CVE reports')
    parser.add_argument('--do-page-report', dest='do_page_report', help='Generate CVE report')
    parser.add_argument('--display-repo-match', dest='display_repo_match', help='Generate artifact map between audits')

    # CVE of the Day
    parser.add_argument('--pass-audit-id', nargs=1, type=int ,dest='command1', help='audit id to insert CveOfTheDay records')
    parser.add_argument('--fetch-cves-of-the-day', dest='fetch_cves_of_the_day', help='Insert CveOfTheDay records for passed audit id')
    parser.add_argument('--populate-cod-historical-data', dest='populate_cod_historical_data', help='populate historical data in CveOfTheDay table')
    # Debug support for CoD
    parser.add_argument('--review-cod-historical-data', dest='review_cod_historical_data', help='Display cumulative cve totals [start_audit_id,stop_audit_is,content_type]')
    parser.add_argument('--dump-cod-historical-data', dest='dump_cod_historical_data', help='Dump cumulative cve totals [start_audit_id,stop_audit_is,content_type]')
    parser.add_argument('--dump-cod-audit-data', dest='dump_cod_audit_data', help='Dump CoD changes for a given audit')
    parser.add_argument('--delete-cves-of-the-day', dest='cve_cod_delete', help='Delete the existing CoD content, set base audit ID')

    parser.add_argument('--call-scan-cve-sla', nargs=2, type=int, dest='AuditIdsforSlacall', help='audit id range for sla function call')
    parser.add_argument('--cve-sla-mode', dest='cve_sla_mode', help='Extract summary SLAs for "develop"')
    parser.add_argument('--cve-sla-apply-audit', dest='cve_sla_apply_audit', help='Apply CVE Age records to an audit')
    parser.add_argument('--cve-sla-apply-test', dest='cve_sla_apply_audit_test', help='Test CVE Age records agains an audit')
    parser.add_argument('--cve-sla-update-all', dest='cve_sla_update_all', help='Update SLA info and recent audits [default|first_id#,last_id#]')
    # Debug support for SLA
    parser.add_argument('--scan-cve-sla', '-A', nargs=1, type=int, dest='AuditIdforSla', help='audit id for sla') #
    parser.add_argument('--scan-cve-sla-old', '-B', nargs=1, type=int, dest='AuditIdforSlaOld', help='Old audit id for sla, uses indexes') #
    parser.add_argument('--cve-sla-dump', action='store_const', const='cve_sla_dump', dest='command', help='Dump the existing CveSla content to a text file')
    parser.add_argument('--cve-sla-delete', action='store_const', const='cve_sla_delete', dest='command', help='Delete the existing CveSla content')

    parser.add_argument('--merge-ingest', nargs=1, type=int, dest='merge_ingest', help='Merge ingest data into audit') #

    parser.add_argument('--api-cve-table', dest='api_cve_table', help='CVE table for "api_cve_list" API')
    parser.add_argument('--refresh-cve-nist', action='store_const', const='refresh_cve_nist', dest='command', help='Refresh from NIST the CVE data')
    parser.add_argument('--cve-link-update', dest='cve_link_update', help='Update CVE links')
    parser.add_argument('--copy-srt-db-latest', action='store_const', const='copy_srt_db_latest', dest='command', help='Copy the latest SRTool for Linux DB')

    parser.add_argument('--check-wrlinux-cve', dest='check_wrlinux_cve', help='Check WR Linux CVE database for CVE(s)')
    parser.add_argument('--skip-db-copy', '-S', action='store_true', dest='skip_db_copy', help='Skip the download of the WR Linux CVE source database')
    parser.add_argument('--skip-to-cve', '-C', dest='skip_to_cve', help='Skip refresh to this one CVE')

    # Common parameters
    parser.add_argument('--audit-id', '-I', dest='audit_id', help='ID of Audit record, if any')
    parser.add_argument('--user-id', '-U', dest='user_id', help='ID of user, if any')
    parser.add_argument('--progress', action='store_true', dest='do_progress', help='Progress output')
    parser.add_argument('--srt-lx-remote', dest='srt_lx_remote', help='remote ULR for SRTool Linux database, if any')
    parser.add_argument('--srt-lx-path', dest='srt_lx_path', help='Path to SRTool Linux database')

    # Repair and checks
    parser.add_argument('--fix-upper-severity', action='store_const', const='fix_upper_severity', dest='command', help='Fix records with upper case severities')
    parser.add_argument('--analyze-cveage-match', dest='analyze_cveage_match', help='analyze CVE Age records mis-match against an audit')

    # Debugging support
    parser.add_argument('--force', '-f', action='store_true', dest='force_update', help='Force update')
    parser.add_argument('--test', '-t', action='store_true', dest='test', help='Test, dry-run')
    parser.add_argument('--count', dest='count', help='Debugging: short run record count')
    parser.add_argument('--skip', dest='skip', help='Debugging: skip record count')
    parser.add_argument('--verbose', '-v', action='store_true', dest='verbose', help='Verbose debugging')
    args = parser.parse_args()

#    _log("WR_STUDIO:ARGS:%s" % str(args))
##    master_log = open("./update_logs/master_log.txt", "a")

    if args.audit_id:
        audit_id = int(args.audit_id)
    else:
        # Explicitly set "0" for "No Audit"
        audit_id = 0

    verbose = args.verbose
    test = args.test
    force_update = args.force_update
    if None != args.count:
        cmd_count = int(args.count)
    if None != args.skip:
        cmd_skip = int(args.skip)
    skip_to_cve = ''
    if args.skip_to_cve:
        skip_to_cve = args.skip_to_cve

    # Connection to SRTool for WR Linux database
    if args.srt_lx_remote:
        srt_lx_remote = args.srt_lx_remote
    else:
        srt_lx_remote = os.environ.get('SRT_LX_REMOTE')
    if args.srt_lx_path:
        srt_lx_path = args.srt_lx_path
    else:
        srt_lx_path = os.environ.get('SRT_LX_PATH')

    progress_set_on(args.do_progress)
    progress_set_max(1)     # Default max, in case job progress is not internally used

    if args.user_id:
        user_id = int(args.user_id)
    else:
        user_id = None

    if args.do_page_report:
        ret = do_page_report(audit_id, args.do_page_report, user_id)
        return(ret)
    elif args.display_repo_match:
        ret = display_repo_match(args.display_repo_match)
        return(ret)

    elif args.fetch_cves_of_the_day:
        fetch_cves_of_the_day(args.fetch_cves_of_the_day)
        ret = 1
    elif args.populate_cod_historical_data:
        print('passed audit_id is for historical CoD data is,',args.populate_cod_historical_data)
        populate_cod_historical_data(args.populate_cod_historical_data)
        ret = 1
    elif args.dump_cod_historical_data:
        args.dump_cod_historical_data += '#'    # In case skip list not appended
        data = args.dump_cod_historical_data.split('#')
        ret = dump_cod_historical_data(data[0],data[1])
    elif args.dump_cod_audit_data:
        ret = dump_cod_audit_data(args.dump_cod_audit_data)

    elif args.review_cod_historical_data:
        args.review_cod_historical_data += '#'    # In case skip list not appended
        data = args.review_cod_historical_data.split('#')
        ret = review_cod_historical_data(data[0],data[1])

    elif args.cve_cod_delete:
        ret = cve_cod_delete(args.cve_cod_delete)

    elif args.AuditIdforSla:
        print('id passed thru command for sla--->',args.AuditIdforSla[0])
        scan_cve_sla(args.AuditIdforSla[0])
        ret = 1
    elif args.AuditIdforSlaOld:
        print('NAMED:id passed thru command for sla--->',args.AuditIdforSlaOld[0])
        scan_cve_sla_idx(args.AuditIdforSlaOld[0])
        ret = 1
    elif args.AuditIdsforSlacall:
        print('ids range passed thru cmd for calling sla function-->',args.AuditIdsforSlacall[0],args.AuditIdsforSlacall[1])
        call_scan_cve_sla(args.AuditIdsforSlacall[0],args.AuditIdsforSlacall[1])
        ret = 1
    elif args.cve_sla_mode:
        cve_sla_mode(int(args.cve_sla_mode))
        ret = 1
    elif args.cve_sla_apply_audit:
        ret = cve_sla_apply_audit(args.cve_sla_apply_audit,is_dry_run=False)
    elif args.cve_sla_apply_audit_test:
        ret = cve_sla_apply_audit(args.cve_sla_apply_audit_test,is_dry_run=True)

    elif args.cve_sla_update_all:
        ret = cve_sla_update_all(args.cve_sla_update_all)

    elif args.api_cve_table:
        api_cve_table(args.api_cve_table)
        ret = 1
    elif 'refresh_cve_nist' == args.command:
        ret = refresh_cve_nist(args.skip_db_copy,skip_to_cve)
    elif args.cve_link_update:
        ret = cve_link_update(args.cve_link_update)
    elif 'copy_srt_db_latest' == args.command:
        print(f"SRT_DB_LATEST: {import_srtool_db(SRTOOL_DEST_DB_PATH)} => {SRTOOL_DEST_DB_PATH}")
        ret = 1

    elif args.check_wrlinux_cve:
        check_wrlinux_cve(args.check_wrlinux_cve)
        ret = 1

    elif 'fix_upper_severity' == args.command:
        ret = fix_upper_severity()
    elif args.analyze_cveage_match:
        ret = analyze_cveage_match(args.analyze_cveage_match)
    elif 'cve_sla_delete' == args.command:
        ret = cve_sla_delete()
    elif 'cve_sla_dump' == args.command:
        ret = cve_sla_dump()

    elif args.merge_ingest:
        ret = merge_ingest(args.merge_ingest[0])

    else:
        print("srtool_cves.py:Command not found")
        ret = 1

    progress_done('Done')
    return(ret)


if __name__ == '__main__':
    srtool_basepath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(sys.argv[0]))))
    exit( main(sys.argv[1:]) )
